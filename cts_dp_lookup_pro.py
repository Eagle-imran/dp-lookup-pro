import asyncio
import datetime
import io
import json
import math
import os
import re
import sys
import time
from typing import Any, Callable, Dict, List, Optional
from xml.sax.saxutils import escape as _xml_escape

import ezdxf
import httpx
import qrcode
from ezdxf.enums import TextEntityAlignment
from openpyxl import Workbook, load_workbook
from PIL import Image, ImageDraw
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import Image as RLImage
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

SERVER_URL = "https://agsmaps.mcgm.gov.in/server/rest/services/Development_Plan_2034/MapServer"
SATELLITE_URL = "https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile"

CACHE_FILENAME = ".cache_store.json"

# 30 days.
#
# Chosen to match how slowly this data actually moves rather than as a generic
# "safe" default. Layer 13 (parcel geometry and area) carries LAST_EDITED_DATE
# 2019-01-23 on every parcel checked - the boundaries are a frozen snapshot.
# Zones, reservations and CRZ tiers are statutory and change only by gazetted
# order. The one genuinely moving part is DP modifications (layer 192), which
# appear sporadically - Byculla 1605 gained one dated 22.08.2024.
#
# The trade: a newly gazetted modification can go unseen for up to 30 days.
# Mitigated by reporting cache age on every hit (see metadata.cache_age_days)
# and by --no-cache for same-day certainty.
DEFAULT_CACHE_TTL_SECONDS = 30 * 24 * 60 * 60

# In-memory tile cache is process-scoped and safe to share across lookups.
_TILE_CACHE: Dict[str, bytes] = {}

# One lookup store per output directory. The store used to be a single module
# global bound to a hardcoded './output/.cache_store.json', which ignored the
# caller's output_dir and wrote relative to the process CWD.
_STORES: Dict[str, Dict[str, Any]] = {}


def cache_path_for(output_dir: str) -> str:
    """Cache store lives beside the bundles it describes, not beside the CWD."""
    return os.path.join(output_dir, CACHE_FILENAME)


def load_disk_cache(output_dir: str) -> Dict[str, Any]:
    path = cache_path_for(output_dir)
    if path in _STORES:
        return _STORES[path]
    store: Dict[str, Any] = {}
    if os.path.exists(path):
        try:
            with open(path, "r", encoding="utf-8") as f:
                loaded = json.load(f)
            if isinstance(loaded, dict):
                store = loaded
        except (OSError, ValueError):
            store = {}
    _STORES[path] = store
    return store


def save_disk_cache(output_dir: str) -> None:
    path = cache_path_for(output_dir)
    store = _STORES.get(path, {})
    try:
        parent = os.path.dirname(path)
        if parent:
            os.makedirs(parent, exist_ok=True)
        with open(path, "w", encoding="utf-8") as f:
            json.dump(store, f, indent=2)
    except OSError as exc:
        # Surface rather than swallow: a read-only or full disk silently
        # disabling the cache used to be invisible.
        print(f"[dp-lookup-pro] WARNING: could not write cache store {path}: {exc}",
              file=sys.stderr)


def bundle_is_intact(result: Dict[str, Any]) -> bool:
    """
    True when every file the cached result promises still exists on disk.

    A cache hit used to hand back paths without checking them, so deleting an
    output folder produced a confident "success" pointing at seven missing files.
    The Excel register is excluded: it is shared, regenerated on demand, and its
    absence does not invalidate the plot bundle.
    """
    files = result.get("export_files")
    if not isinstance(files, dict):
        return False
    for key, path in files.items():
        if key == "master_excel_register" or not isinstance(path, str):
            continue
        if not os.path.exists(path):
            return False
    return True


def read_cache_entry(output_dir: str, key: str, ttl_seconds: int) -> Optional[Dict[str, Any]]:
    """
    Return a cached result, or None when it is absent, expired, in the legacy
    format, or no longer backed by files on disk.
    """
    entry = load_disk_cache(output_dir).get(key)
    if not isinstance(entry, dict):
        return None
    # Legacy entries were the bare result dict with no timestamp. They predate
    # the CRZ and road fixes, so treat them as expired rather than trusting them.
    cached_at = entry.get("cached_at")
    result = entry.get("result")
    if not isinstance(cached_at, (int, float)) or not isinstance(result, dict):
        return None
    if ttl_seconds >= 0 and (time.time() - cached_at) > ttl_seconds:
        return None
    if not bundle_is_intact(result):
        return None

    result = json.loads(json.dumps(result))
    age = max(0.0, time.time() - cached_at)
    meta = result.setdefault("metadata", {})
    meta["cached_at"] = datetime.datetime.fromtimestamp(cached_at).strftime("%Y-%m-%d %H:%M:%S")
    meta["cache_age_days"] = round(age / 86400.0, 2)
    meta["cache_expires_in_days"] = (
        None if ttl_seconds < 0 else round(max(0.0, ttl_seconds - age) / 86400.0, 2)
    )
    return result


def write_cache_entry(output_dir: str, key: str, result: Dict[str, Any]) -> None:
    store = load_disk_cache(output_dir)
    store[key] = {"cached_at": time.time(), "result": result}
    save_disk_cache(output_dir)


# --- Input validation -------------------------------------------------------
# village and cts_number are interpolated into an ArcGIS WHERE clause. Restrict
# them to the characters real MCGM values actually use, then double any quote
# that survives, so a value like  1' OR '1'='1  cannot alter the query.
_VILLAGE_RE = re.compile(r"^[A-Za-z0-9 ()./-]{1,64}$")
_CTS_RE = re.compile(r"^[A-Za-z0-9 ()./-]{1,32}$")


def sanitize_query_value(value: Any, pattern: "re.Pattern[str]", label: str, max_len: int) -> str:
    text = str(value).strip()
    if not text:
        raise ValueError(f"{label} must not be empty")
    if len(text) > max_len:
        raise ValueError(f"{label} is too long (max {max_len} characters)")
    if not pattern.match(text):
        raise ValueError(
            f"{label} contains unsupported characters. "
            f"Allowed: letters, digits, space, and ( ) . / -"
        )
    return text.replace("'", "''")


def _esc(value: Any) -> str:
    """Escape a data value for embedding in XML/KML or ReportLab markup."""
    return _xml_escape("" if value is None else str(value))


def usable_json(resp: Any) -> Optional[Dict[str, Any]]:
    """
    Parsed JSON body, or None when the response is unusable.

    ArcGIS reports failures as HTTP 200 carrying {"error": {...}}. Reading
    .get("features", []) on such a body yields an empty list that is
    indistinguishable from a genuine no-match, which is how three separate
    silent failures reached production. Every parse goes through here.
    """
    if isinstance(resp, Exception) or resp is None:
        return None
    if getattr(resp, "status_code", None) != 200:
        return None
    try:
        body = resp.json()
    except (ValueError, AttributeError):
        return None
    if not isinstance(body, dict) or "error" in body:
        return None
    return body

def latlon_to_tile(lat: float, lon: float, zoom: int):
    n = 2 ** zoom
    x_tile = int((lon + 180.0) / 360.0 * n)
    lat_rad = math.radians(lat)
    y_tile = int((1.0 - math.asinh(math.tan(lat_rad)) / math.pi) / 2.0 * n)
    return x_tile, y_tile

def tile_to_latlon(x_tile: int, y_tile: int, zoom: int):
    n = 2 ** zoom
    lon = x_tile / n * 360.0 - 180.0
    lat_rad = math.atan(math.sinh(math.pi * (1 - 2 * y_tile / n)))
    lat = math.degrees(lat_rad)
    return lat, lon

async def fetch_tile_cached(client: httpx.AsyncClient, zoom: int, ty: int, tx: int) -> bytes:
    key = f"{zoom}/{ty}/{tx}"
    if key in _TILE_CACHE:
        return _TILE_CACHE[key]
    url = f"{SATELLITE_URL}/{zoom}/{ty}/{tx}"
    try:
        r = await client.get(url)
        if r.status_code == 200:
            _TILE_CACHE[key] = r.content
            return r.content
    except Exception:
        pass
    return b""

def export_geojson(wgs_rings: list, properties: dict, output_path: str):
    feature = {
        "type": "Feature",
        "geometry": {
            "type": "Polygon" if len(wgs_rings) == 1 else "MultiPolygon",
            "coordinates": wgs_rings if len(wgs_rings) == 1 else [wgs_rings]
        },
        "properties": properties
    }
    geojson_data = {
        "type": "FeatureCollection",
        "name": f"CTS_{properties.get('cts_no')}_{properties.get('village')}",
        "crs": {
            "type": "name",
            "properties": {"name": "urn:ogc:def:crs:OGC:1.3:CRS84"}
        },
        "features": [feature]
    }
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(geojson_data, f, indent=2)

def wgs84_to_utm43n(lon: float, lat: float) -> tuple:
    """
    Converts WGS84 (lon, lat) in degrees to UTM Zone 43N (Easting, Northing) in meters.
    WGS84 Ellipsoid: a = 6378137.0 m, f = 1/298.257223563
    Central Meridian = 75.0° E (UTM Zone 43N covers Mumbai 72.8° E).
    """
    a = 6378137.0
    f = 1.0 / 298.257223563
    b = a * (1.0 - f)
    e2 = (a**2 - b**2) / (a**2)
    ep2 = (a**2 - b**2) / (b**2)
    
    k0 = 0.9996
    lon0 = 75.0
    
    lat_rad = math.radians(lat)
    lon_rad = math.radians(lon)
    lon0_rad = math.radians(lon0)
    
    N = a / math.sqrt(1.0 - e2 * math.sin(lat_rad)**2)
    T = math.tan(lat_rad)**2
    C = ep2 * math.cos(lat_rad)**2
    A = (lon_rad - lon0_rad) * math.cos(lat_rad)
    
    M = a * (
        (1.0 - e2/4.0 - 3.0*e2**2/64.0 - 5.0*e2**3/256.0) * lat_rad
        - (3.0*e2/8.0 + 3.0*e2**2/32.0 + 45.0*e2**3/1024.0) * math.sin(2.0 * lat_rad)
        + (15.0*e2**2/256.0 + 45.0*e2**3/1024.0) * math.sin(4.0 * lat_rad)
        - (35.0*e2**3/3072.0) * math.sin(6.0 * lat_rad)
    )
    
    x = k0 * N * (
        A
        + (1.0 - T + C) * A**3 / 6.0
        + (5.0 - 18.0*T + T**2 + 72.0*C - 58.0*ep2) * A**5 / 120.0
    ) + 500000.0
    
    y = k0 * (
        M + N * math.tan(lat_rad) * (
            A**2 / 2.0
            + (5.0 - T + 9.0*C + 4.0*C**2) * A**4 / 24.0
            + (61.0 - 58.0*T + T**2 + 600.0*C - 330.0*ep2) * A**6 / 720.0
        )
    )
    return x, y

def polygon_signed_area(ring: list) -> float:
    """Shoelace. Positive = counter-clockwise winding."""
    n = len(ring)
    return sum(
        ring[i][0] * ring[(i + 1) % n][1] - ring[(i + 1) % n][0] * ring[i][1]
        for i in range(n)
    ) / 2.0


def offset_polygon_inward(ring: list, distance: float) -> list:
    """
    True parallel inward offset of a closed polygon. Returns a LIST of rings.

    This is what a building setback actually is: every point on the result is
    exactly `distance` from the nearest boundary edge.

    Two earlier approaches were wrong and both were caught by measurement:

      1. Pulling each vertex radially toward the centroid by
         min(distance, dist*0.3). Neither perpendicular nor the requested
         distance - on WORLI 733 the "3 m" line sat 1.15-3.00 m from the
         boundary.
      2. A hand-rolled miter offset. Exact on convex plots, but it
         self-intersects on concave ones, and real CTS parcels are concave
         (WORLI 733 has 6 reflex corners). It measured 0.68 m against a 3 m
         target.

    Shapely's buffer resolves self-intersection properly, which is the whole
    difficulty here. A setback can legitimately split a plot into more than one
    buildable island, so every resulting ring is returned.

    Returns [] when the plot cannot sustain the offset, so the caller omits the
    line rather than drawing a misleading one.
    """
    from shapely.geometry import Polygon
    from shapely.geometry.base import BaseMultipartGeometry

    pts = list(ring)
    if len(pts) > 1 and math.isclose(pts[0][0], pts[-1][0], abs_tol=1e-9) \
            and math.isclose(pts[0][1], pts[-1][1], abs_tol=1e-9):
        pts = pts[:-1]
    if len(pts) < 3 or distance <= 0:
        return []

    try:
        poly = Polygon(pts)
        if not poly.is_valid:
            poly = poly.buffer(0)          # repair self-touching input
        shrunk = poly.buffer(-abs(distance), join_style=2)  # 2 = mitre
    except Exception:
        return []

    if shrunk.is_empty:
        return []

    parts = list(shrunk.geoms) if isinstance(shrunk, BaseMultipartGeometry) else [shrunk]
    out = []
    for part in parts:
        exterior = getattr(part, "exterior", None)
        if exterior is None:
            continue
        coords = [(x, y) for x, y in exterior.coords]
        if len(coords) > 1 and coords[0] == coords[-1]:
            coords = coords[:-1]
        if len(coords) >= 3 and abs(polygon_signed_area(coords)) >= 1.0:
            out.append(coords)
    return out


# --- DXF construction --------------------------------------------------------
# Layer table and the local projection, lifted out of export_dxf (420 lines).

# name -> (ACI colour, linetype). Order is the order they appear in the legend.
# name, ACI colour, linetype, lineweight.
#
# Lineweight is in 1/100 mm and must be one of AutoCAD's discrete values. Every
# layer previously carried the default (-3), so a plotted sheet rendered the
# metric grid at the same weight as the plot boundary and read completely flat.
# The boundary is the heaviest line on the sheet, setbacks sit below it, and
# reference material (grid, neighbours, annotation) sits below that.
DXF_LAYERS = (
    ("0_GRID_AXIS", 8, "DASHED", 9),
    ("C-PLOT-BDY", 1, None, 50),
    ("C-PROP-HATCH", 252, None, 9),
    ("C-SETBACK-3M", 3, "DASHED", 25),
    ("C-SETBACK-6M", 70, "DASHED", 25),
    ("C-ADJN-PLOTS", 2, None, 13),
    ("C-ROAD-ALIGN", 4, "DASHED", 35),
    ("C-RESTRICT-ZONE", 6, "PHANTOM", 35),
    ("C-ANNO-TEXT", 7, None, 13),
    ("C-ANNO-DIMS", 5, None, 13),
    ("C-NORTH-ARROW", 7, None, 18),
    ("C-TITLE-BLOCK", 4, None, 18),
)

# Degrees to metres at Mumbai's latitude. Longitude scales with cos(lat);
# latitude is effectively constant over a parcel.
METRES_PER_DEG_LAT = 111132.0
METRES_PER_DEG_LON_EQUATOR = 111319.5


def new_dxf_document():
    """R2010 document in metric units with the project's layer table."""
    doc = ezdxf.new("R2010")
    doc.header["$INSUNITS"] = 6      # metres
    doc.header["$MEASUREMENT"] = 1   # ISO metric
    try:
        doc.linetypes.add("DASHED", pattern="A, 1.0, -0.5")
        doc.linetypes.add("PHANTOM", pattern="A, 2.0, -0.5, 0.5, -0.5, 0.5, -0.5")
    except Exception:
        # A document that already defines them is fine; the layers below still work.
        pass
    for name, colour, linetype, lineweight in DXF_LAYERS:
        attrs = {"name": name, "color": colour, "lineweight": lineweight}
        if linetype:
            attrs["linetype"] = linetype
        doc.layers.add(**attrs)
    return doc


def project_rings_to_local_metres(wgs_rings: list) -> Dict[str, Any]:
    """
    WGS84 degrees to local metres with the plot centroid at (0, 0).

    An architect wants 1 CAD unit = 1 metre and the plot at the origin, not
    UTM coordinates in the millions. Accurate to well under a centimetre over a
    parcel, which is finer than MCGM's own digitisation.
    """
    r0 = wgs_rings[0]
    lon0 = sum(p[0] for p in r0) / len(r0)
    lat0 = sum(p[1] for p in r0) / len(r0)
    mpd_lon = METRES_PER_DEG_LON_EQUATOR * math.cos(math.radians(lat0))
    mpd_lat = METRES_PER_DEG_LAT

    return {
        "lon0": lon0, "lat0": lat0,
        "mpd_lon": mpd_lon, "mpd_lat": mpd_lat,
        "local_rings": [[((p[0] - lon0) * mpd_lon, (p[1] - lat0) * mpd_lat) for p in ring]
                        for ring in wgs_rings],
    }


def to_local_metres(points: list, lon0: float, lat0: float,
                    mpd_lon: float, mpd_lat: float) -> list:
    """Project an arbitrary WGS84 ring/path into the same local frame."""
    return [((p[0] - lon0) * mpd_lon, (p[1] - lat0) * mpd_lat) for p in points]


def clip_segment_to_box(p1: tuple, p2: tuple, x_lo: float, y_lo: float,
                        x_hi: float, y_hi: float) -> Optional[tuple]:
    """
    The portion of segment p1-p2 that lies inside the box, or None.

    Liang-Barsky. Clipping rather than keeping whole segments matters because MCGM
    road centrelines run kilometres long; retaining an outside endpoint would drag
    the sheet border out with it.
    """
    x1, y1 = float(p1[0]), float(p1[1])
    x2, y2 = float(p2[0]), float(p2[1])
    dx, dy = x2 - x1, y2 - y1
    t0, t1 = 0.0, 1.0
    for p, q in ((-dx, x1 - x_lo), (dx, x_hi - x1), (-dy, y1 - y_lo), (dy, y_hi - y1)):
        if p == 0:
            if q < 0:
                return None          # parallel to this edge and outside it
        else:
            r = q / p
            if p < 0:
                if r > t1:
                    return None
                t0 = max(t0, r)
            else:
                if r < t0:
                    return None
                t1 = min(t1, r)
    return ((x1 + t0 * dx, y1 + t0 * dy), (x1 + t1 * dx, y1 + t1 * dy))


def clip_path_to_window(path: list, x_lo: float, y_lo: float,
                        x_hi: float, y_hi: float) -> List[list]:
    """
    Contiguous runs of `path` clipped to the window.

    A segment is kept when it CROSSES the window, not merely when one of its
    vertices sits inside it. MCGM returns road centrelines whose vertices can be a
    kilometre apart, so testing vertices alone dropped roads that ran straight past
    the plot - AMBIVALI 807 has a named 27.4 m frontage and an empty
    C-ROAD-ALIGN layer because of it.
    """
    runs: List[list] = []
    run: list = []
    for i in range(len(path) - 1):
        seg = clip_segment_to_box(path[i], path[i + 1], x_lo, y_lo, x_hi, y_hi)
        if seg is None:
            if run:
                runs.append(run)
                run = []
            continue
        start, end = seg
        if run and abs(run[-1][0] - start[0]) < 1e-9 and abs(run[-1][1] - start[1]) < 1e-9:
            run.append(end)
        else:
            if run:
                runs.append(run)
            run = [start, end]
    if run:
        runs.append(run)
    return runs


def text_extents(entity) -> Optional[tuple]:
    """
    (x0, y0, x1, y1) of a TEXT entity as it will actually render.

    Placement arithmetic cannot tell you whether a label overruns its frame or
    collides with another label - only the rendered width can, and that depends
    on the font. Every text fault found when WORLI 733 was opened in AutoCAD was
    invisible to coordinate checks for exactly this reason: legend rows escaping
    the panel, the UTM tie-in lying across a dimension, two grid labels stacked
    at the corner.

    Rotation is applied to the corners so rotated dimension labels measure
    correctly. Returns None when the font engine is unavailable, so callers must
    treat measurement as best-effort rather than guaranteed.
    """
    try:
        from ezdxf.tools.text_size import text_size
        size = text_size(entity)
        w, h = size.width, size.cap_height
    except Exception:
        try:
            h = float(entity.dxf.height)
            w = len(entity.dxf.text) * h * 0.72
        except Exception:
            return None
    try:
        x, y = float(entity.dxf.insert[0]), float(entity.dxf.insert[1])
    except Exception:
        return None
    rot = math.radians(float(getattr(entity.dxf, "rotation", 0.0) or 0.0))
    if not rot:
        return (x, y, x + w, y + h)
    cos_r, sin_r = math.cos(rot), math.sin(rot)
    pts = [(x + dx * cos_r - dy * sin_r, y + dx * sin_r + dy * cos_r)
           for dx, dy in ((0.0, 0.0), (w, 0.0), (w, h), (0.0, h))]
    xs = [p[0] for p in pts]
    ys = [p[1] for p in pts]
    return (min(xs), min(ys), max(xs), max(ys))


def boxes_overlap(a: Optional[tuple], b: Optional[tuple], slack: float = 0.0) -> bool:
    """True when two (x0, y0, x1, y1) boxes intersect. Unmeasurable text never collides."""
    if a is None or b is None:
        return False
    return not (a[2] + slack < b[0] or b[2] + slack < a[0]
                or a[3] + slack < b[1] or b[3] + slack < a[1])


def nudge_text_clear(entity, obstacles: list, step: float,
                     direction: tuple = (0.0, 1.0), limit: int = 10) -> bool:
    """
    Shift a label along `direction` until it stops overlapping anything in `obstacles`.

    Adjoining-parcel labels sit at the parcel centroid and boundary dimensions at
    the edge midpoints; on WORLI 733 'CTS 733A' landed squarely on the '10.23m'
    dimension, and on AMBIVALI 807 the '6.20m' and '3.47m' dimensions of two short
    adjacent edges landed on each other. Neither placement is wrong in isolation,
    so the label that moves is the one added later. Dimensions push outward along
    the edge normal so they stay associated with their edge.

    Returns True once clear, False if it never cleared within `limit` steps.
    """
    if step <= 0:
        return True
    dx, dy = direction
    for _ in range(limit + 1):
        box = text_extents(entity)
        if box is None:
            return True
        if not any(boxes_overlap(box, other) for other in obstacles):
            return True
        ins = entity.dxf.insert
        entity.set_placement((float(ins[0]) + dx * step, float(ins[1]) + dy * step))
    return False


# One source of truth for the legend rows. This list used to live inside the
# drawing function while the sheet-sizing code carried its own hardcoded count of
# 12, so the two could drift apart silently.
DXF_LEGEND_ROWS = (
    ('C-PLOT-BDY',      'PLOT BOUNDARY - gross plot area'),
    ('C-PROP-HATCH',    'Gross plot area (fill)'),
    ('C-ROAD-ALIGN',    'Abutting road alignment / frontage'),
    ('C-SETBACK-3M',    '3.0 m setback line (true parallel offset)'),
    ('C-SETBACK-6M',    '6.0 m setback line (true parallel offset)'),
    # This layer carries advisory text only. CRZ status comes from a point
    # identify, not a polygon, so there is no zone boundary to draw and the row
    # must not imply one.
    ('C-RESTRICT-ZONE', 'CRZ / Metro restriction (advisory note)'),
    ('C-ADJN-PLOTS',    'Adjoining CTS plots'),
    ('C-ANNO-DIMS',     'Boundary segment dimensions (m)'),
    ('C-ANNO-TEXT',     'Plot metadata'),
    ('C-NORTH-ARROW',   'True north'),
    ('0_GRID_AXIS',     'Metric grid, 0,0 at plot centroid'),
    ('C-TITLE-BLOCK',   'Sheet border, legend, title block'),
)

# Rows of vertical padding the panel needs beyond its content rows: the heading
# gaps above LAYER LEGEND and PLOT DATA, and the margin below the last row.
LEGEND_PAD_ROWS = 4.15


def dxf_title_block_lines(properties: Dict[str, Any],
                          utm_cx: float, utm_cy: float) -> List[str]:
    """
    The sheet's title block.

    The Property Card caution is here rather than in PLOT DATA because it applies
    to the drawing as a whole. MCGM's record and MCGM's own digitised polygon are
    already both printed in the panel; the Property Card is a third source this
    tool cannot read, and on WORLI 733 the owner measured it 5-7% away from the
    MCGM figure.
    """
    return [
        "MCGM DEVELOPMENT PLAN 2034 - CAD BASE",
        f"PLOT: CTS {properties.get('cts_no')} ({properties.get('village')})",
        "PURPOSE: Architectural concept & massing base",
        "SCALE: 1:1 METRIC (1 CAD unit = 1 metre)",
        "ORIGIN: Plot centroid (0.00, 0.00)",
        f"UTM 43N CENTROID: E {utm_cx:.2f} / N {utm_cy:.2f}",
        "SETBACKS ARE INDICATIVE - confirm against DCPR 2034",
        "PLOT AREA PER PROPERTY CARD MAY DIFFER FROM THE MCGM",
        "RECORD - reconcile both before any FSI calculation",
    ]


def title_block_height(lg_row: float, n_lines: int) -> float:
    """Height the title block needs for `n_lines` of MTEXT plus its top margin."""
    return lg_row * (n_lines + 1.2)


def legend_column_height(lg_row: float, n_legend: int, n_data: int) -> float:
    """
    Height the legend panel needs to actually contain its rows.

    Was `lg_row * (n_legend + 9.5)`, which never accounted for the PLOT DATA
    rows at all - the panel border cut across the last three entries on every
    drawing ever generated.
    """
    return lg_row * (n_legend + n_data + LEGEND_PAD_ROWS)


# Longest value a PLOT DATA row carries before it wraps onto a continuation line.
LEGEND_VALUE_MAX_CHARS = 34


def legend_data_row(label: str, value: Any) -> List[str]:
    """
    One PLOT DATA row, wrapped onto continuation lines when the value is long.

    AMBIVALI 807's frontage is 'Jay Prakash Road Part II Dadabhai Road to Versova
    Metro.' - long enough to push that single row 19.8 m past the panel edge and
    16.7 m outside the sheet border. Widening the panel to fit it would make the
    panel wider than the plot, so the value wraps instead.
    """
    text = str(value)
    if len(text) <= LEGEND_VALUE_MAX_CHARS:
        return [f"{label:<16}: {text}"]
    lines, current = [], ""
    for word in text.split():
        if current and len(current) + 1 + len(word) > LEGEND_VALUE_MAX_CHARS:
            lines.append(current)
            current = word
        else:
            current = f"{current} {word}".strip()
    if current:
        lines.append(current)
    return [f"{label:<16}: {lines[0]}"] + [f"{'':<16}  {ln}" for ln in lines[1:]]


def legend_data_rows(properties: Dict[str, Any], setback_status: Dict[float, bool],
                     measured_area_sqm: Optional[float] = None) -> List[str]:
    """
    The PLOT DATA rows.

    Both areas are printed on purpose. MCGM's approved record, MCGM's own
    digitised parcel polygon and the Property Card are three independent sources
    that do not agree: on WORLI 733 the record and the polygon differ by 0.30%,
    but on AMBIVALI 807 they differ by 6.10% - 123 sq m. Printing a single figure
    invites an FSI calculation off a number nobody reconciled, so the sheet shows
    the record, shows what the drawn boundary actually measures, and names the
    gap.
    """
    sb3 = "drawn" if setback_status.get(3.0) else "NOT VIABLE - plot too narrow"
    sb6 = "drawn" if setback_status.get(6.0) else "NOT VIABLE - plot too narrow"
    area_note = properties.get('area_source') or ''
    derived = 'derived' in area_note
    record_area = properties.get('area_sqm')

    rows = legend_data_row("GROSS PLOT AREA", f"{record_area} sq m")
    rows += legend_data_row(
        "AREA SOURCE", "DERIVED from boundary" if derived else "MCGM approved record")
    if measured_area_sqm:
        delta = ""
        if not derived and isinstance(record_area, (int, float)) and record_area:
            pct = (measured_area_sqm - record_area) / record_area * 100.0
            delta = f"  ({pct:+.2f}% vs record)"
        rows += legend_data_row("MEASURED (BDY)", f"{measured_area_sqm:.2f} sq m{delta}")
    rows += legend_data_row(
        "CTS / VILLAGE", f"{properties.get('cts_no')} / {properties.get('village')}")
    rows += legend_data_row(
        "WARD / ZONE", f"{properties.get('ward')} / {properties.get('zone')}")
    rows += legend_data_row(
        "ABUTTING ROAD",
        f"{properties.get('abutting_road')} ({properties.get('road_width')})")
    # The drawing carries a short 'CRZ: ...' marker because a sentence does not fit
    # on a small plot; the instruction that goes with it belongs here.
    crz_value = properties.get('crz_buffer_flag')
    if str(crz_value or '').upper().startswith('YES'):
        crz_value = f"{crz_value} - verify with MCGM/MCZMA"
    rows += legend_data_row("CRZ STATUS", crz_value)
    rows += legend_data_row("METRO BUFFER", properties.get('metro_buffer_flag'))
    rows += legend_data_row("3.0 m SETBACK", sb3)
    rows += legend_data_row("6.0 m SETBACK", sb6)
    return rows


def draw_dxf_legend_column(msp, properties: Dict[str, Any], setback_status: Dict[float, bool],
                           utm_cx: float, utm_cy: float,
                           lg_x0: float, lg_w: float, b_max_y: float, b_min_y: float,
                           char_h: float, dim_char_h: float, scale: float,
                           data_rows: Optional[List[str]] = None) -> float:
    """
    Legend, plot-data panel and title block, stacked down the right-hand column.

    Legend swatches are drawn ON their own layers, so each sample line carries
    that layer's real colour and linetype - the swatch is the layer rather than
    a picture of one.

    Both panel frames are drawn LAST and sized from the measured extents of the
    text inside them. Sizing them first from a row count is what let four legend
    rows overrun the panel edge and three fall outside it entirely. Returns the
    rightmost x reached, so the caller can put the sheet border outside it.
    """
    if data_rows is None:
        data_rows = legend_data_rows(properties, setback_status)

    lg_row = max(char_h * 1.9, 2.2)
    left = lg_x0 + lg_row * 0.4
    lg_y1 = b_max_y - scale * 0.05
    right_edge = lg_x0 + lg_w
    emitted = []

    def place(text: str, height: float, x: float, y: float):
        entity = msp.add_text(
            text, dxfattribs={'layer': 'C-TITLE-BLOCK', 'height': height})
        entity.set_placement((x, y))
        emitted.append(entity)
        return entity

    # ---- LEGEND ----------------------------------------------------------
    y = lg_y1 - lg_row * 1.2
    place("LAYER LEGEND", char_h * 0.95, left, y)
    y -= lg_row * 1.3

    swatch_w = lg_row * 1.5
    for layer_name, meaning in DXF_LEGEND_ROWS:
        # sample line drawn ON its own layer, so it carries that layer's colour
        # and linetype - the swatch is the layer, not a picture of it.
        msp.add_line((left, y + lg_row * 0.22),
                     (left + swatch_w, y + lg_row * 0.22),
                     dxfattribs={'layer': layer_name})
        place(f"{layer_name}  -  {meaning}", dim_char_h * 0.92,
              left + swatch_w + lg_row * 0.5, y)
        y -= lg_row

    # ---- PLOT DATA (what an architect needs before massing) --------------
    y -= lg_row * 0.6
    place("PLOT DATA", char_h * 0.95, left, y)
    y -= lg_row * 1.25

    for row in data_rows:
        place(row, dim_char_h * 0.92, left, y)
        y -= lg_row

    # Frame sized to what is actually inside it, in both axes.
    measured_right = right_edge
    for entity in emitted:
        box = text_extents(entity)
        if box:
            measured_right = max(measured_right, box[2])
    lg_x1 = measured_right + lg_row * 0.4
    lg_y0 = min(y + lg_row - lg_row * 0.8,
                lg_y1 - legend_column_height(lg_row, len(DXF_LEGEND_ROWS), len(data_rows)))
    msp.add_lwpolyline(
        [(lg_x0, lg_y0), (lg_x1, lg_y0), (lg_x1, lg_y1), (lg_x0, lg_y1)],
        dxfattribs={'layer': 'C-TITLE-BLOCK', 'closed': True})

    # ---- TITLE BLOCK (bottom of the legend column) -----------------------
    title_lines = dxf_title_block_lines(properties, utm_cx, utm_cy)
    tb_char_h = dim_char_h * 0.92
    tb_h = title_block_height(lg_row, len(title_lines))
    tb_y0 = b_min_y + scale * 0.05
    tb_x1 = lg_x1
    mtext = msp.add_mtext(
        "\n".join(title_lines),
        dxfattribs={'layer': 'C-TITLE-BLOCK', 'char_height': tb_char_h})
    mtext.set_location((left, tb_y0 + tb_h - lg_row * 0.5))
    msp.add_lwpolyline(
        [(lg_x0, tb_y0), (tb_x1, tb_y0), (tb_x1, tb_y0 + tb_h), (lg_x0, tb_y0 + tb_h)],
        dxfattribs={'layer': 'C-TITLE-BLOCK', 'closed': True})

    return lg_x1




def export_dxf(wgs_rings: list, properties: dict, output_path: str,
               neighbors: Optional[list] = None, roads: Optional[list] = None):
    """
    Generates a scale-accurate Multi-Layered AutoCAD DXF CAD drawing file (.dxf)
    centered around Local Origin (0, 0) in Real-World Metric Meters (1 CAD Unit = 1 Meter).
    Configures active modelspace viewport & extents for instant framing in AutoCAD 2024.
    
    Layers:
    - 0_GRID_AXIS: Metric Axis Grid & Axis Labels
    - C-PLOT-BDY: Primary CTS Plot Boundary Polyline (Closed, Red)
    - C-PROP-HATCH: Solid Plot Fill Pattern
    - C-SETBACK-3M: 3.0m Concept Setback Line for Massing (Dashed Green)
    - C-SETBACK-6M: 6.0m Tower Setback Line (Dashed Dark Green)
    - C-ADJN-PLOTS: Adjoining CTS Plot Boundaries & CTS Labels (Yellow)
    - C-ROAD-ALIGN: Abutting Road Alignment Vector (Cyan)
    - C-RESTRICT-ZONE: Metro Rail / CRZ Restriction Zone (Phantom Magenta)
    - C-ANNO-TEXT: Plot Centroid Attribute Block & Parcel Metadata
    - C-ANNO-DIMS: Automated Boundary Segment Dimensions in Meters
    - C-TITLE-BLOCK: Architectural Sheet Border Frame & Title Block Box
    """
    doc = new_dxf_document()
    msp = doc.modelspace()
    
    # 1. Fast Scale-Accurate Metric Projection around Local Origin (0, 0)
    proj = project_rings_to_local_metres(wgs_rings)
    lon0, lat0 = proj["lon0"], proj["lat0"]
    meters_per_deg_lon, meters_per_deg_lat = proj["mpd_lon"], proj["mpd_lat"]
    local_rings = proj["local_rings"]
    
    lr0 = local_rings[0]
    all_x = [p[0] for p in lr0]
    all_y = [p[1] for p in lr0]
    min_x, max_x = min(all_x), max(all_x)
    min_y, max_y = min(all_y), max(all_y)
    
    # Calculate real-world UTM Zone 43N centroid for title block metadata
    utm_cx, utm_cy = wgs84_to_utm43n(lon0, lat0)
    
    cx, cy = 0.0, 0.0
    width = max_x - min_x
    height = max_y - min_y
    scale = max(width, height, 30.0)
    char_h = max(1.2, scale * 0.035)
    dim_char_h = max(0.9, scale * 0.025)
    
    # 2. 0_GRID_AXIS: Metric Axis Grid around (0, 0)
    grid_step = 50.0 if scale > 80 else 25.0
    g_min_x = math.floor((min_x - scale * 0.4) / grid_step) * grid_step
    g_max_x = math.ceil((max_x + scale * 0.4) / grid_step) * grid_step
    g_min_y = math.floor((min_y - scale * 0.4) / grid_step) * grid_step
    g_max_y = math.ceil((max_y + scale * 0.4) / grid_step) * grid_step
    
    gx = g_min_x
    while gx <= g_max_x:
        msp.add_line((gx, g_min_y), (gx, g_max_y), dxfattribs={'layer': '0_GRID_AXIS'})
        msp.add_text(f"{int(gx)}m", dxfattribs={'layer': '0_GRID_AXIS', 'height': dim_char_h * 0.8}).set_placement((gx + 0.5, g_min_y + 0.5))
        gx += grid_step
        
    gy = g_min_y
    while gy <= g_max_y:
        msp.add_line((g_min_x, gy), (g_max_x, gy), dxfattribs={'layer': '0_GRID_AXIS'})
        # The bottom-left corner is where the X label row and the Y label column
        # meet: both emitted a label at exactly the same point, so the two sat
        # stacked on top of each other. The corner's Y value is redundant there.
        if gy > g_min_y:
            msp.add_text(f"{int(gy)}m", dxfattribs={'layer': '0_GRID_AXIS', 'height': dim_char_h * 0.8}).set_placement((g_min_x + 0.5, gy + 0.5))
        gy += grid_step

    # 3. C-PLOT-BDY: Primary Plot Boundary (Closed Polyline)
    for loc_ring in local_rings:
        poly = msp.add_lwpolyline(loc_ring, dxfattribs={'layer': 'C-PLOT-BDY', 'closed': True})
        poly.dxf.const_width = max(0.15, scale * 0.005)
        
        # Plot fill. Deliberately transparent: as an opaque solid it hid anything
        # an architect placed underneath - a survey underlay, a satellite image,
        # an imported DP extract - which is most of what this drawing is for.
        try:
            hatch = msp.add_hatch(color=252, dxfattribs={'layer': 'C-PROP-HATCH'})
            hatch.paths.add_polyline_path(loc_ring, is_closed=True)
            hatch.transparency = 0.65
        except Exception:
            pass

    # 4. C-SETBACK-3M & C-SETBACK-6M: true parallel setback lines.
    #
    # Each is a genuine perpendicular offset from every boundary edge (miter
    # joins at the corners), so an architect can build massing directly to these
    # lines. If the plot is too small to sustain an offset the line is omitted
    # rather than drawn wrong, and that is recorded for the legend.
    setback_status = {}
    for spec_layer, spec_dist in (('C-SETBACK-3M', 3.0), ('C-SETBACK-6M', 6.0)):
        drawn = 0
        for loc_ring in local_rings:
            for offset_ring in offset_polygon_inward(loc_ring, spec_dist):
                msp.add_lwpolyline(offset_ring, dxfattribs={'layer': spec_layer, 'closed': True})
                drawn += 1
        setback_status[spec_dist] = drawn > 0
        if not drawn:
            # A marker, not a sentence. The full wording lives in the PLOT DATA
            # panel; spelled out here it ran 3.8x the width of an 8 m plot, which
            # is exactly the size of plot that triggers this message.
            msp.add_text(
                f"{spec_dist:.1f}m SETBACK N/A",
                dxfattribs={'layer': spec_layer, 'height': dim_char_h}
            ).set_placement((min_x, min_y - dim_char_h * (2.5 if spec_dist == 3.0 else 4.0)))

    # 5. C-ANNO-DIMS: Automated Boundary Side Dimension Lines in Meters
    #
    # Two short adjacent edges put their labels close enough to overlap - AMBIVALI
    # 807's '6.20m' and '3.47m' landed on each other. Each label is pushed further
    # out along its own edge normal until it clears the ones already placed, so it
    # stays associated with the edge it measures.
    placed_boxes = []
    for loc_ring in local_rings:
        n_pts = len(loc_ring)
        for i in range(n_pts):
            p1 = loc_ring[i]
            p2 = loc_ring[(i + 1) % n_pts]
            dx = p2[0] - p1[0]
            dy = p2[1] - p1[1]
            seg_len = math.sqrt(dx*dx + dy*dy)
            # Sub-metre slivers are digitisation noise and their labels overlap
            # into an unreadable cluster; the segment itself is still drawn.
            if seg_len > 1.0:
                mx = (p1[0] + p2[0]) / 2.0
                my = (p1[1] + p2[1]) / 2.0
                # Outward normal
                nx = -dy / seg_len
                ny = dx / seg_len
                if (mx + nx - cx)**2 + (my + ny - cy)**2 < (mx - cx)**2 + (my - cy)**2:
                    nx = -nx
                    ny = -ny
                tx = mx + nx * (dim_char_h * 1.8)
                ty = my + ny * (dim_char_h * 1.8)
                angle_deg = math.degrees(math.atan2(dy, dx))
                if angle_deg > 90 or angle_deg < -90:
                    angle_deg += 180
                
                txt_elem = msp.add_text(f"{seg_len:.2f}m", dxfattribs={
                    'layer': 'C-ANNO-DIMS',
                    'height': dim_char_h,
                    'rotation': angle_deg
                })
                txt_elem.set_placement((tx, ty))
                nudge_text_clear(txt_elem, placed_boxes, dim_char_h * 1.1,
                                 direction=(nx, ny), limit=4)
                box = text_extents(txt_elem)
                if box:
                    placed_boxes.append(box)

    # 6. C-ADJN-PLOTS: Adjoining CTS Plot Boundaries
    if neighbors:
        for n in neighbors:
            n_rings = n.get('rings', [])
            n_cts = n.get('cts_no', 'N/A')
            for n_ring in n_rings:
                n_loc_ring = [((pt[0] - lon0) * meters_per_deg_lon, (pt[1] - lat0) * meters_per_deg_lat) for pt in n_ring]
                msp.add_lwpolyline(n_loc_ring, dxfattribs={'layer': 'C-ADJN-PLOTS', 'closed': True})
                if n_loc_ring:
                    ncx = sum(p[0] for p in n_loc_ring) / len(n_loc_ring)
                    ncy = sum(p[1] for p in n_loc_ring) / len(n_loc_ring)
                    label = msp.add_text(f"CTS {n_cts}", dxfattribs={
                        'layer': 'C-ADJN-PLOTS',
                        'height': dim_char_h * 0.9
                    })
                    label.set_placement((ncx, ncy))
                    # A neighbour's centroid can coincide with one of our own
                    # boundary dimensions ('CTS 733A' over '10.23m' on WORLI 733).
                    # The dimension is load-bearing for the architect; the
                    # neighbour label is reference, so the label moves.
                    nudge_text_clear(label, placed_boxes, dim_char_h * 1.2, limit=6)
                    box = text_extents(label)
                    if box:
                        placed_boxes.append(box)

    # 6b. C-ROAD-ALIGN: abutting road centreline(s).
    # This layer was declared but never populated - an architect had no idea
    # which edge was the frontage, which is what governs the front setback.
    # MCGM returns whole road networks - one polyline came back 3.8 km long with
    # 2472 vertices. Clip to the plot's vicinity so the sheet stays legible and
    # the drawing extents stay sane.
    clip = scale * 0.9
    cx_lo, cx_hi = min_x - clip, max_x + clip
    cy_lo, cy_hi = min_y - clip, max_y + clip

    road_drawn = 0
    for road_ring in (roads or []):
        loc = [((pt[0] - lon0) * meters_per_deg_lon, (pt[1] - lat0) * meters_per_deg_lat)
               for pt in road_ring]
        for seg in clip_path_to_window(loc, cx_lo, cy_lo, cx_hi, cy_hi):
            if len(seg) >= 2:
                msp.add_lwpolyline(seg, dxfattribs={'layer': 'C-ROAD-ALIGN'})
                road_drawn += 1

    if road_drawn:
        # Shortened on purpose. AMBIVALI 807's frontage name runs to 56
        # characters, which on the drawing is wider than the plot itself; the
        # full name and width are in the PLOT DATA panel.
        road_label = str(properties.get('abutting_road') or 'N/A')
        if len(road_label) > 24:
            road_label = road_label[:23].rstrip(" ,.") + "..."
        road_text = msp.add_text(
            f"ABUTTING ROAD: {road_label} ({properties.get('road_width')})",
            dxfattribs={'layer': 'C-ROAD-ALIGN', 'height': dim_char_h})
        road_text.set_placement((min_x, min_y - dim_char_h * 1.2))
        # Sits below the plot, so it clears downward - pushing it up would move it
        # onto the geometry it is annotating.
        nudge_text_clear(road_text, placed_boxes, dim_char_h * 1.3,
                         direction=(0.0, -1.0), limit=6)
        box = text_extents(road_text)
        if box:
            placed_boxes.append(box)

    # 7. C-RESTRICT-ZONE: CRZ and Metro development restrictions.
    # Markers, not sentences. Spelled out in full these ran 3.3x the width of an
    # 8 m plot; the full wording is in the PLOT DATA panel, which has room for it.
    restrict_notes = []
    crz_flag = str(properties.get("crz_buffer_flag") or "")
    if crz_flag.upper().startswith("YES"):
        restrict_notes.append(f"CRZ: {crz_flag}")
        restrict_notes.append("DEVELOPMENT RESTRICTED")
    # Prefix test, not equality. The lookup sets this to "YES (Metro Buffer Zone)",
    # so `== "YES"` never matched and the metro restriction was silently absent
    # from every drawing ever generated - AMBIVALI 807 is in a metro buffer and its
    # DXF said nothing. Same failure mode as the CRZ check above, which is why both
    # now share the shape.
    if str(properties.get("metro_buffer_flag") or "").upper().startswith("YES"):
        msp.add_circle((cx, cy), radius=scale * 0.6, dxfattribs={'layer': 'C-RESTRICT-ZONE'})
        restrict_notes.append("METRO RAIL BUFFER")
    # Stacked above the plot, clear of the dimension band. The dimension labels
    # sit dim_char_h * 1.8 outside the boundary plus their own cap height, so a
    # note placed nearer than that lands on top of the top-edge dimensions - which
    # is what put 'PLOT CENTROID' across '3.30m' on WORLI 733 and across two
    # dimensions on AMBIVALI 807.
    # Notes are the last annotation placed, so they clear everything already on
    # the sheet and then stack upward from wherever each one landed. Placing them
    # at fixed offsets put 'Development restricted...' straight through the
    # 'CTS 732A' neighbour label on WORLI 733.
    note_y = max_y + dim_char_h * 4.4
    for note in restrict_notes:
        note_text = msp.add_text(note, dxfattribs={
            'layer': 'C-RESTRICT-ZONE', 'height': dim_char_h})
        note_text.set_placement((min_x, note_y))
        nudge_text_clear(note_text, placed_boxes, dim_char_h * 1.2, limit=14)
        box = text_extents(note_text)
        if box:
            placed_boxes.append(box)
            note_y = box[3] + dim_char_h * 1.1
        else:
            note_y += dim_char_h * 2.1

    # 8. C-ANNO-TEXT: plot identity only, at the centroid.
    #
    # This used to print an eight-line metadata block across the middle of the
    # plot - covering the boundary and both setback lines, i.e. exactly the area
    # an architect needs clear to draw in. Rendering the DXF made that obvious
    # in a way no geometry check could. The full metadata already lives in the
    # PLOT DATA panel beside the legend, so only the identifier stays here.
    msp.add_text(
        f"CTS {properties.get('cts_no')}",
        dxfattribs={'layer': 'C-ANNO-TEXT', 'height': char_h}
    ).set_placement((cx, cy), align=TextEntityAlignment.MIDDLE_CENTER)

    # The UTM tie-in used to be printed here as a 56-character string above the
    # plot. It was the widest annotation on every sheet (0.91x the plot width on
    # AMBIVALI 807, ~6x on a small plot), it collided with the top-edge
    # dimensions, and it is already in the title block verbatim as
    # "UTM 43N CENTROID". Carrying it twice bought nothing but the collision.

    # 8b. C-NORTH-ARROW: orientation. Local +Y is true north because the metric
    # projection maps latitude straight onto Y.
    na_x = max_x + scale * 0.16
    na_y = max_y - scale * 0.06
    na_r = scale * 0.07
    msp.add_circle((na_x, na_y), radius=na_r, dxfattribs={'layer': 'C-NORTH-ARROW'})
    msp.add_lwpolyline(
        [(na_x, na_y + na_r * 1.5), (na_x - na_r * 0.5, na_y - na_r * 0.7), (na_x + na_r * 0.5, na_y - na_r * 0.7)],
        dxfattribs={'layer': 'C-NORTH-ARROW', 'closed': True})
    msp.add_text("N", dxfattribs={'layer': 'C-NORTH-ARROW', 'height': char_h}) \
        .set_placement((na_x - char_h * 0.35, na_y + na_r * 1.8))

    # 9. C-TITLE-BLOCK: sheet border, legend panel and title block.
    #
    # The border must be sized from EVERYTHING drawn, not from the plot alone.
    # Roads and adjoining parcels routinely extend well beyond the plot, so a
    # border derived from min_x/max_x left them outside the frame and put the
    # legend directly on top of them.
    content_x, content_y = [], []
    for entity in msp:
        try:
            if entity.dxftype() == 'LWPOLYLINE':
                for vx, vy in entity.get_points('xy'):
                    content_x.append(vx)
                    content_y.append(vy)
            elif entity.dxftype() == 'LINE':
                content_x += [entity.dxf.start.x, entity.dxf.end.x]
                content_y += [entity.dxf.start.y, entity.dxf.end.y]
            elif entity.dxftype() == 'CIRCLE':
                ctr, rad = entity.dxf.center, entity.dxf.radius
                content_x += [ctr.x - rad, ctr.x + rad]
                content_y += [ctr.y - rad, ctr.y + rad]
            elif entity.dxftype() == 'TEXT':
                # Text was invisible to this scan, which is precisely why labels
                # ended up outside the border - the thing sizing the border could
                # not see them.
                box = text_extents(entity)
                if box:
                    content_x += [box[0], box[2]]
                    content_y += [box[1], box[3]]
        except Exception:
            continue
    if not content_x:
        content_x, content_y = [min_x, max_x], [min_y, max_y]

    pad = max(scale * 0.10, 3.0)
    b_min_x = min(content_x) - pad
    plot_right = max(content_x) + pad
    b_min_y = min(content_y) - pad
    b_max_y = max(content_y) + pad

    # What the drawn boundary actually measures, which is not always what MCGM's
    # record says: AMBIVALI 807 differs by 123 sq m (6.10%). Both go on the sheet.
    measured_area = sum(abs(polygon_signed_area(r)) for r in local_rings) or None
    data_rows = legend_data_rows(properties, setback_status, measured_area)

    # The legend column carries the legend rows, the plot-data rows and the title
    # block. On a small plot that stack is taller than the drawing, so grow the
    # sheet downward rather than letting the legend run off the bottom. Both
    # heights are derived from the row lists themselves - the old figures were
    # magic constants that had never counted the PLOT DATA rows at all.
    _lg_row = max(char_h * 1.9, 2.2)
    _needed = (legend_column_height(_lg_row, len(DXF_LEGEND_ROWS), len(data_rows))
               + title_block_height(_lg_row, len(dxf_title_block_lines(properties, utm_cx, utm_cy)))
               + pad * 3)
    if (b_max_y - b_min_y) < _needed:
        b_min_y = b_max_y - _needed

    # Legend panel sits to the right of all geometry, so it never covers anything.
    lg_w = max(scale * 0.95, 26.0)
    lg_x0 = plot_right + scale * 0.06

    # The legend is drawn BEFORE the border and reports how far right it actually
    # reached, so the border can contain it. Sized the other way round, four
    # legend rows overran the panel and one escaped the sheet entirely.
    lg_x1 = draw_dxf_legend_column(
        msp, properties, setback_status, utm_cx, utm_cy,
        lg_x0, lg_w, b_max_y, b_min_y, char_h, dim_char_h, scale,
        data_rows=data_rows,
    )
    b_max_x = max(lg_x0 + lg_w, lg_x1) + scale * 0.06

    msp.add_lwpolyline(
        [(b_min_x, b_min_y), (b_max_x, b_min_y), (b_max_x, b_max_y), (b_min_x, b_max_y)],
        dxfattribs={'layer': 'C-TITLE-BLOCK', 'closed': True})

    # Set Header Extents & Active Modelspace Viewport Zoom Framing for AutoCAD 2024
    doc.header['$EXTMIN'] = (b_min_x, b_min_y, 0)
    doc.header['$EXTMAX'] = (b_max_x, b_max_y, 0)
    doc.header['$LIMMIN'] = (b_min_x, b_min_y)
    doc.header['$LIMMAX'] = (b_max_x, b_max_y)
    
    try:
        doc.set_modelspace_vport(
            height=(b_max_y - b_min_y) * 1.08,
            center=((b_min_x + b_max_x) / 2.0, (b_min_y + b_max_y) / 2.0),
        )
    except Exception:
        pass
        
    doc.saveas(output_path)

def export_kml(wgs_rings: list, properties: dict, output_path: str):
    ring0 = wgs_rings[0]
    coords_str = " ".join(f"{p[0]},{p[1]},0" for p in ring0)
    
    # Every interpolated value is XML-escaped. An unescaped & or < in a village or
    # road name produced a KML that Google Earth refused to parse.
    kml_content = f"""<?xml version="1.0" encoding="UTF-8"?>
<kml xmlns="http://www.opengis.net/kml/2.2">
  <Document>
    <name>CTS {_esc(properties.get('cts_no'))} ({_esc(properties.get('village'))}) DP Remark</name>
    <description>MCGM DP 2034 Spatial Polygon Export</description>
    <Style id="plotStyle">
      <LineStyle>
        <color>ff00e6ff</color>
        <width>4</width>
      </LineStyle>
      <PolyStyle>
        <color>4000e6ff</color>
      </PolyStyle>
    </Style>
    <Placemark>
      <name>CTS {_esc(properties.get('cts_no'))} - {_esc(properties.get('village'))}</name>
      <styleUrl>#plotStyle</styleUrl>
      <ExtendedData>
        <Data name="Ward"><value>{_esc(properties.get('ward'))}</value></Data>
        <Data name="Zone"><value>{_esc(properties.get('zone'))}</value></Data>
        <Data name="Area_sqm"><value>{_esc(properties.get('area_sqm'))}</value></Data>
        <Data name="Status"><value>{_esc(properties.get('status_badge'))}</value></Data>
        <Data name="CRZ_Status"><value>{_esc(properties.get('crz_buffer_flag'))}</value></Data>
        <Data name="Metro_Buffer"><value>{_esc(properties.get('metro_buffer_flag'))}</value></Data>
        <Data name="Abutting_Road"><value>{_esc(properties.get('abutting_road'))} ({_esc(properties.get('road_width'))})</value></Data>
      </ExtendedData>
      <Polygon>
        <extrude>1</extrude>
        <altitudeMode>clampToGround</altitudeMode>
        <outerBoundaryIs>
          <LinearRing>
            <coordinates>{coords_str}</coordinates>
          </LinearRing>
        </outerBoundaryIs>
      </Polygon>
    </Placemark>
  </Document>
</kml>"""
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(kml_content)

def build_pdf_doc(pdf_path, status_badge, status_summary, village, attrs, cts_number, zone, des_desc, des_code, mod_approval, mod_label, crz_buffer_flag, metro_buffer_flag, road_name, road_width, dp_snapshot_path, qr_bytes, map_link, sat_snapshot_path, neighbors, area_source_label="Approved cadastral area (MCGM record)"):
    doc = SimpleDocTemplate(pdf_path, pagesize=letter, leftMargin=36, rightMargin=36, topMargin=36, bottomMargin=36)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=15, leading=19, textColor=colors.HexColor('#1A237E'))
    subtitle_style = ParagraphStyle('Sub', parent=styles['Normal'], fontSize=9, leading=12, textColor=colors.HexColor('#424242'))
    body_style = ParagraphStyle('Body', parent=styles['Normal'], fontSize=8.5, leading=11)
    bold_style = ParagraphStyle('Bold', parent=styles['Normal'], fontSize=8.5, leading=11, fontName='Helvetica-Bold')

    story = []
    story.append(Paragraph("<b>MCGM DEVELOPMENT PLAN 2034 — SPATIAL REMARK DOCKET</b>", title_style))
    now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    story.append(Paragraph(f"Official Spatial Query Report | Page 1 of 2 | Generated: {now_str}", subtitle_style))
    story.append(Spacer(1, 6))

    banner_color = colors.HexColor('#E8F5E9') if 'CLEAR' in status_badge else (colors.HexColor('#FFFDE7') if 'MODIFIED' in status_badge else colors.HexColor('#FFEBEE'))
    banner_text = f"<b>STATUS: {_esc(status_badge)}</b> — {_esc(status_summary)}"
    banner_table = Table([[Paragraph(banner_text, ParagraphStyle('BText', parent=body_style, fontSize=9.5, leading=12))]], colWidths=[540])
    banner_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), banner_color),
        ('PADDING', (0,0), (-1,-1), 5),
        ('BOX', (0,0), (-1,-1), 1, colors.HexColor('#BDBDBD')),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
    ]))
    story.append(banner_table)
    story.append(Spacer(1, 6))

    # ReportLab Paragraph parses a mini-XML markup language, so every DATA value is
    # escaped while the intentional <b>/<a> tags in the templates are left alone.
    # An unescaped & or < in a road name or modification label previously raised
    # during doc.build(), crashing the run after all network work had succeeded.
    mod_label_trimmed = mod_label[:60] + ('...' if len(mod_label) > 60 else '')
    table_data = [
        [Paragraph("<b>Attribute</b>", bold_style), Paragraph("<b>Details</b>", bold_style), Paragraph("<b>Planning Remarks</b>", bold_style)],
        [Paragraph("Village / Ward", body_style), Paragraph(f"{_esc(village.upper())} (Ward {_esc(attrs['WARD'])})", body_style), Paragraph("MCGM Administrative Division", body_style)],
        [Paragraph("Plot CTS No.", body_style), Paragraph(f"CTS {_esc(cts_number)} ({_esc(attrs['TYPE'])})", body_style), Paragraph("City Survey Cadastral Parcel", body_style)],
        [Paragraph("Plot Area", body_style), Paragraph(f"{_esc(attrs['AREA_APP_SQ_MTRS'])} sq m" if attrs['AREA_APP_SQ_MTRS'] else "Not on record", body_style), Paragraph(_esc(area_source_label), body_style)],
        [Paragraph("Land-Use Zone", body_style), Paragraph(f"<b>Zone {_esc(zone)}</b>", body_style), Paragraph("Primary Zoning Classification", body_style)],
        [Paragraph("PLU Designation", body_style), Paragraph(f"{_esc(des_desc)} ({_esc(des_code)})", body_style), Paragraph("Existing Amenity Designation", body_style)],
        [Paragraph("DP Modification", body_style), Paragraph(f"Approval: {_esc(mod_approval)}", body_style), Paragraph(_esc(mod_label_trimmed), body_style)],
        [Paragraph("CRZ Status", body_style), Paragraph(f"<b>{_esc(crz_buffer_flag)}</b>", body_style), Paragraph("Coastal Regulation Zone Layer Query", body_style)],
        [Paragraph("Metro Buffer", body_style), Paragraph(_esc(metro_buffer_flag), body_style), Paragraph("Layer 1550 Metro Rail Influence", body_style)],
        [Paragraph("Abutting Road", body_style), Paragraph(f"<b>{_esc(road_name)}</b> ({_esc(road_width)})", body_style), Paragraph("DP 2034 Road Access &amp; Width", body_style)],
    ]
    
    t = Table(table_data, colWidths=[110, 215, 215])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#ECEFF1')),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CFD8DC')),
        ('PADDING', (0,0), (-1,-1), 2.5),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
    ]))
    story.append(t)
    story.append(Spacer(1, 6))

    dp_img_rl = RLImage(dp_snapshot_path, width=360, height=360)
    qr_img_rl = RLImage(qr_bytes, width=100, height=100)
    
    qr_cell = [
        Paragraph("<b>Scan for Live Interactive Map:</b>", body_style),
        Spacer(1, 4),
        qr_img_rl,
        Spacer(1, 4),
        Paragraph(f"<a href='{_esc(map_link)}'>Open ArcGIS Web Map</a>", ParagraphStyle('Link', parent=body_style, textColor=colors.blue, fontSize=8))
    ]
    
    media_table = Table([[dp_img_rl, qr_cell]], colWidths=[370, 170])
    media_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ALIGN', (1,0), (1,0), 'CENTER'),
    ]))
    story.append(media_table)

    story.append(PageBreak())
    story.append(Paragraph("<b>HIGH-RESOLUTION SATELLITE & ADJOINING PLOT CLUSTER ANALYSIS</b>", title_style))
    story.append(Paragraph(f"Official Spatial Query Report | Page 2 of 2 | Plot CTS {_esc(cts_number)} ({_esc(village.upper())})", subtitle_style))
    story.append(Spacer(1, 8))

    sat_img_rl = RLImage(sat_snapshot_path, width=540, height=340)
    story.append(sat_img_rl)
    story.append(Spacer(1, 10))

    story.append(Paragraph("<b>Adjoining / Neighboring CTS Parcels (Spatial Cluster):</b>", ParagraphStyle('H2', parent=bold_style, fontSize=10, textColor=colors.HexColor('#1A237E'))))
    story.append(Spacer(1, 4))

    adj_table_data = [[Paragraph("<b>Adjoining CTS No.</b>", bold_style), Paragraph("<b>Village</b>", bold_style), Paragraph("<b>Plot Area (sq m)</b>", bold_style)]]
    for n in neighbors[:8]:
        adj_table_data.append([
            Paragraph(f"CTS {_esc(n['cts_no'])}", body_style),
            Paragraph(_esc(n['village'].upper()), body_style),
            Paragraph(f"{_esc(n['area_sqm'])} sq m" if n['area_sqm'] != 'N/A' else "N/A", body_style)
        ])
    if not neighbors:
        adj_table_data.append([Paragraph("No immediate adjoining CTS polygons detected in buffer", body_style), Paragraph("-", body_style), Paragraph("-", body_style)])

    adj_table = Table(adj_table_data, colWidths=[180, 180, 180])
    adj_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#ECEFF1')),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CFD8DC')),
        ('PADDING', (0,0), (-1,-1), 3.5),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
    ]))
    story.append(adj_table)

    doc.build(story)

# --- pure helpers ------------------------------------------------------------
# Extracted from lookup_plot_pro, which had grown to 730 lines doing network IO,
# geometry, parsing, rendering, export and logging in one scope. Everything below
# is free of IO and shared state, so it is unit-testable without touching MCGM.

WEB_MERCATOR_R = 20037508.342789244


def mercator_to_wgs84(x: float, y: float, precision: int = 7) -> List[float]:
    """Web Mercator (EPSG:102100/3857) metres to WGS84 degrees."""
    lon = round((x / WEB_MERCATOR_R) * 180, precision)
    lat = round(
        (math.atan(math.exp(y / WEB_MERCATOR_R * math.pi)) * 2 - math.pi / 2) * 180 / math.pi,
        precision,
    )
    return [lon, lat]


def resolve_plot_area(attrs: Dict[str, Any]) -> Dict[str, Any]:
    """
    Decide which area to report, and say which one it is.

    MCGM leaves AREA_APP_SQ_MTRS null on a meaningful number of parcels
    (MALABAR HILL 518, TARDEO 264). SHAPE.AREA - the digitised polygon's own
    area - is populated there and is already in true ground square metres:
    verified to match AREA_APP_SQ_MTRS exactly on plots carrying both.

    It is NOT the same authority. It is the drawn area, not the approved
    cadastral one, and the two differ by up to 7% where both exist. So it is a
    labelled fallback, never a silent substitution.

    Returns {"area_sqm", "area_source", "note"} - note is None when the approved
    figure was available.
    """
    approved = attrs.get("AREA_APP_SQ_MTRS")
    geometry = attrs.get("SHAPE.AREA")
    if approved not in (None, "", 0):
        return {"area_sqm": approved,
                "area_source": "approved (MCGM AREA_APP_SQ_MTRS)",
                "note": None}
    if isinstance(geometry, (int, float)) and geometry > 0:
        return {"area_sqm": round(float(geometry), 2),
                "area_source": "derived from plot geometry - MCGM has no approved area on record",
                "note": ("MCGM has no approved area for this plot; area is derived from the "
                         "digitised boundary and is indicative only")}
    return {"area_sqm": None, "area_source": "unavailable", "note": None}


def prepare_geometry(rings: list) -> Dict[str, Any]:
    """
    Everything derived from the parcel rings: centroid, bounding box, the
    half-extent used for map windows, and the WGS84 reprojection.
    """
    ring = rings[0]
    cx = sum(p[0] for p in ring) / len(ring)
    cy = sum(p[1] for p in ring) / len(ring)

    pts = [p for r in rings for p in r]
    xs = [p[0] for p in pts]
    ys = [p[1] for p in pts]
    mcx, mcy = (min(xs) + max(xs)) / 2, (min(ys) + max(ys)) / 2
    span = max(max(xs) - min(xs), max(ys) - min(ys))

    lon, lat = mercator_to_wgs84(cx, cy, precision=6)
    return {
        "cx": cx, "cy": cy, "mcx": mcx, "mcy": mcy,
        "xs": xs, "ys": ys,
        "d": max(50, span / 2),
        "lat": lat, "lon": lon,
        "wgs_rings": [[mercator_to_wgs84(p[0], p[1]) for p in r] for r in rings],
    }


def bundle_folder_name(village: Any, cts_number: Any) -> str:
    """Filesystem-safe bundle folder, e.g. 'malabar_hill_cts_16-738'."""
    cts_clean = str(cts_number).replace("/", "-").replace("\\", "-")
    village_clean = (str(village).lower().strip()
                     .replace("/", "-").replace("\\", "-").replace(" ", "_"))
    return f"{village_clean}_cts_{cts_clean}"


# MCGM records unnamed carriageway under several placeholder names, and its own
# typo "Exisiting Road" is the commonest.
GENERIC_ROAD_NAMES = ("Exisiting Road", "EXISTING", "Road", "None")


def select_road(road_map: Dict[str, Dict[str, Any]]) -> Dict[str, Any]:
    """
    Pick the abutting road to report, best evidence first:

      1. a real name AND a recorded width
      2. a real name OR a recorded width
      3. whatever was found, so the field is never empty when something matched

    Behaviour-preserving extraction of the selection already in lookup_plot_pro.
    """
    def real_name(r):
        return r.get("name") not in GENERIC_ROAD_NAMES

    def real_width(r):
        return r.get("width") != "N/A"

    values = list(road_map.values())
    named_with_width = [r for r in values if real_width(r) and real_name(r)]
    named_or_width = [r for r in values if real_name(r) or real_width(r)]
    roads = named_with_width or named_or_width or values
    if not roads:
        return {"name": "None", "width": "None"}
    return {"name": roads[0]["name"], "width": roads[0]["width"]}


def crz_flag_from(crz_item: Optional[Dict[str, Any]]) -> str:
    """
    Report the CRZ sub-tier rather than a bare yes.

    The attribute name varies by layer: 14 uses `category` ("II"), 1548 uses
    `Category`, 1264 uses `CLASS` (already "CRZ II").
    """
    if not crz_item:
        return "NO (Outside CRZ Buffer)"
    attrs = crz_item.get("attributes", {}) or {}
    tier = None
    for key in ("category", "Category", "CATEGORY", "CLASS", "Class"):
        val = attrs.get(key)
        if val and str(val).strip().lower() not in ("", "null", "none"):
            tier = str(val).strip()
            break
    if not tier:
        return f"YES ({crz_item.get('layerName')})"
    label = tier if tier.upper().startswith("CRZ") else f"CRZ {tier}"
    return f"YES ({label})"


def derive_status(mod_item, des_item, rv_item, des_desc, des_code,
                  res_code, res_type, mod_approval) -> Dict[str, str]:
    """Planning status badge and one-line summary, in precedence order."""
    if mod_item:
        return {"badge": "\U0001F7E1 MODIFIED (DP Notification Order)",
                "summary": f"Modified via {mod_approval}"}
    if des_item and des_desc != "None":
        return {"badge": f"\U0001F534 RESERVED / DESIGNATED ({des_desc})",
                "summary": f"Designated as {des_desc} ({des_code})"}
    if rv_item and res_code != "None":
        return {"badge": f"\U0001F534 RESERVED ({res_type})",
                "summary": f"Reserved under {res_code}"}
    return {"badge": "\U0001F7E2 CLEAR (No Reservation)",
            "summary": "Unreserved Land Parcel"}


# --- ArcGIS request construction ---------------------------------------------
# Layer sets and payload builders, lifted out of lookup_plot_pro so the exact
# parameters are visible, testable, and no longer 300-character single lines.
# Values here mirror the live-validated requests exactly - do not "tidy" them
# without re-verifying against the server.

PLANNING_LAYER_IDS = [0, 46, 47, 192, 1550]      # zone, reservation, designation, DP mod, metro
ROAD_LAYERS = "visible:193,194,44,45,2224"
NEIGHBOUR_LAYERS = "visible:13"

# Probe offsets outward from each boundary edge, in Web Mercator units. Three
# distances because a single nudge missed frontage on plots set back from the
# carriageway.
ROAD_EDGE_NUDGES = (6.0, 15.0, 25.0)
ROAD_EDGE_PROBE_LIMIT = 12


def identify_payload(x: float, y: float, layers: str, mcx: float, mcy: float, d: float,
                     tolerance: int, return_geometry: bool) -> Dict[str, str]:
    """One ArcGIS /identify POST body. mapExtent + imageDisplay set the pixel
    scale that `tolerance` is measured in, so they travel together."""
    return {
        "geometry": f"{x},{y}",
        "geometryType": "esriGeometryPoint",
        "sr": "102100",
        "layers": layers,
        "tolerance": str(tolerance),
        "mapExtent": f"{mcx - d},{mcy - d},{mcx + d},{mcy + d}",
        "imageDisplay": "1000,1000,96",
        "returnGeometry": "true" if return_geometry else "false",
        "f": "json",
    }


def planning_layers(crz_layer_ids: list) -> str:
    ids = ",".join(str(i) for i in PLANNING_LAYER_IDS + list(crz_layer_ids))
    return f"visible:{ids}"


def map_export_params(x0: float, y0: float, x1: float, y1: float,
                      width: int = 1000, height: int = 1000, dpi: int = 144) -> Dict[str, str]:
    """The /export call. Consistently the slowest request in a lookup."""
    return {
        "bbox": f"{x0},{y0},{x1},{y1}",
        "bboxSR": "102100",
        "imageSR": "102100",
        "size": f"{width},{height}",
        "format": "png",
        "transparent": "false",
        "dpi": str(dpi),
        "f": "image",
    }


def road_probe_points(outer_ring: list, mcx: float, mcy: float,
                      xs: list, ys: list) -> List[tuple]:
    """
    Where to look for the abutting road.

    Two earlier versions were wrong. Polygon /query against layers 193/194 never
    worked - those layers reject spatial queries and answer HTTP 200 with an
    error body that read as "no roads found". Probing only the centroid and two
    bounding-box corners missed frontage entirely, because on an irregular
    parcel those corners fall outside the polygon.

    Now: centroid and both corners (kept - on some parcels a corner lands nearer
    the frontage than any edge midpoint, and dropping them regressed WORLI 947),
    plus the midpoints of the longest edges pushed outward at three distances.
    Additive by design; road_map de-duplicates.
    """
    edge_probes = []
    for i in range(len(outer_ring)):
        p1 = outer_ring[i]
        p2 = outer_ring[(i + 1) % len(outer_ring)]
        ex, ey = p2[0] - p1[0], p2[1] - p1[1]
        seg_len = math.hypot(ex, ey)
        if seg_len <= 0.01:
            continue
        emx, emy = (p1[0] + p2[0]) / 2.0, (p1[1] + p2[1]) / 2.0
        enx, eny = -ey / seg_len, ex / seg_len
        # flip the normal so it points away from the parcel centre
        if (emx + enx - mcx) ** 2 + (emy + eny - mcy) ** 2 < (emx - mcx) ** 2 + (emy - mcy) ** 2:
            enx, eny = -enx, -eny
        for nudge in ROAD_EDGE_NUDGES:
            edge_probes.append((seg_len, (emx + enx * nudge, emy + eny * nudge)))

    edge_probes.sort(key=lambda e: e[0], reverse=True)
    return [
        (mcx, mcy),
        (min(xs), min(ys)),
        (max(xs), max(ys)),
    ] + [pt for _, pt in edge_probes[:ROAD_EDGE_PROBE_LIMIT]]


def neighbour_probe_points(mcx: float, mcy: float, d: float) -> List[tuple]:
    """Four cardinal offsets around the parcel."""
    off = max(15.0, d * 0.7)
    return [(mcx + off, mcy), (mcx - off, mcy), (mcx, mcy + off), (mcx, mcy - off)]


# --- image rendering ---------------------------------------------------------
# Both renderers take bytes and return a PIL image, so they can be exercised
# with synthetic input and never need the network.


def render_dp_map(base_png: Optional[bytes], rings: list, bbox: tuple, size: tuple,
                  labels: Dict[str, Any]) -> "Image.Image":
    """
    DP zoning map with the plot outlined and a legend.

    `base_png` is the server's rendered map, or None when the fetch failed - in
    which case a blank ground is used so the report still builds rather than
    crashing on a missing image.
    """
    x0, y0, x1, y1 = bbox
    width, height = size

    if base_png:
        base = Image.open(io.BytesIO(base_png)).convert("RGBA")
    else:
        base = Image.new("RGBA", (width, height), (240, 240, 240, 255))

    overlay = Image.new("RGBA", base.size, (255, 255, 255, 0))
    draw = ImageDraw.Draw(overlay)

    def to_px(x, y):
        return ((x - x0) / (x1 - x0) * width, (y1 - y) / (y1 - y0) * height)

    for ring in rings:
        draw.polygon([to_px(p[0], p[1]) for p in ring],
                     fill=(255, 23, 68, 45), outline=(255, 23, 68, 255), width=5)

    # north arrow
    nx, ny = width - 80, 80
    draw.ellipse([nx - 30, ny - 30, nx + 30, ny + 30],
                 fill=(255, 255, 255, 230), outline=(0, 0, 0, 255), width=2)
    draw.polygon([(nx, ny - 22), (nx - 10, ny + 12), (nx + 10, ny + 12)], fill=(220, 0, 0, 255))
    draw.text((nx - 5, ny - 48), "N", fill=(0, 0, 0, 255), font_size=24)

    lw, lh = 420, 160
    lx0, ly0 = width - lw - 30, height - lh - 30
    draw.rectangle([lx0, ly0, lx0 + lw, ly0 + lh],
                   fill=(255, 255, 255, 235), outline=(0, 0, 0, 255), width=2)
    draw.rectangle([lx0 + 20, ly0 + 20, lx0 + 50, ly0 + 42],
                   fill=(255, 23, 68, 100), outline=(255, 23, 68, 255), width=2)
    draw.text((lx0 + 60, ly0 + 20), f"Plot Boundary (CTS {labels['cts']})",
              fill=(0, 0, 0, 255), font_size=18)
    draw.text((lx0 + 20, ly0 + 55), f"Village: {labels['village']} | Ward: {labels['ward']}",
              fill=(0, 0, 0, 255), font_size=16)
    draw.text((lx0 + 20, ly0 + 85), f"Zone: {labels['zone']} | Area: {labels['area']} sq m",
              fill=(0, 0, 0, 255), font_size=16)
    draw.text((lx0 + 20, ly0 + 115), f"Status: {labels['status']}",
              fill=(0, 0, 0, 255), font_size=16)

    return Image.alpha_composite(base, overlay)


def stitch_satellite(tiles: list, coords: list, grid_dim: int, wgs_rings: list,
                     bounds: tuple, labels: Dict[str, Any]) -> "Image.Image":
    """
    Esri tiles stitched into one canvas with the plot outlined.

    `bounds` is (top_lat, left_lon, bot_lat, right_lon) for the whole grid.
    Tiles that failed arrive as falsy or as exceptions and are skipped, leaving
    the dark ground showing rather than aborting.
    """
    top_lat, left_lon, bot_lat, right_lon = bounds
    canvas_w = canvas_h = grid_dim * 256
    canvas = Image.new("RGBA", (canvas_w, canvas_h), (40, 40, 40, 255))

    for (gx, gy), blob in zip(coords, tiles):
        if not blob or isinstance(blob, Exception):
            continue
        try:
            canvas.paste(Image.open(io.BytesIO(blob)).convert("RGBA"), (gx * 256, gy * 256))
        except Exception:
            continue

    overlay = Image.new("RGBA", (canvas_w, canvas_h), (255, 255, 255, 0))
    draw = ImageDraw.Draw(overlay)

    def to_px(lon, lat):
        return ((lon - left_lon) / (right_lon - left_lon) * canvas_w,
                (top_lat - lat) / (top_lat - bot_lat) * canvas_h)

    for ring in wgs_rings:
        draw.polygon([to_px(p[0], p[1]) for p in ring],
                     fill=(255, 235, 59, 50), outline=(255, 235, 59, 255), width=5)

    slw, slh = 320, 110
    sx, sy = canvas_w - slw - 15, canvas_h - slh - 15
    draw.rectangle([sx, sy, sx + slw, sy + slh],
                   fill=(0, 0, 0, 210), outline=(255, 255, 255, 255), width=2)
    draw.rectangle([sx + 12, sy + 12, sx + 35, sy + 30],
                   fill=(255, 235, 59, 100), outline=(255, 235, 59, 255), width=2)
    draw.text((sx + 42, sy + 12), f"Satellite Boundary (CTS {labels['cts']})",
              fill=(255, 255, 255, 255), font_size=13)
    draw.text((sx + 12, sy + 38), f"Village: {labels['village']} | Ward: {labels['ward']}",
              fill=(255, 255, 255, 255), font_size=12)
    draw.text((sx + 12, sy + 60), f"Lat: {labels['lat']:.6f} | Lon: {labels['lon']:.6f}",
              fill=(255, 255, 255, 255), font_size=12)
    draw.text((sx + 12, sy + 82), f"Adjoining Parcels Identified: {labels['neighbours']}",
              fill=(255, 255, 255, 255), font_size=12)

    return Image.alpha_composite(canvas, overlay)


async def lookup_plot_pro(
    village: str,
    cts_number: str,
    output_dir: str = "./output",
    use_cache: bool = True,
    cache_ttl_seconds: int = DEFAULT_CACHE_TTL_SECONDS,
    timeout_seconds: float = 20.0,
    on_data: Optional[Callable[[Dict[str, Any]], None]] = None,
) -> Dict[str, Any]:
    """
    Ultra-Fast Enterprise DP Plot Lookup Pro Tool (AutoCAD DXF + GeoJSON + KML Exports).

    use_cache          - set False to force a fresh network lookup.
    cache_ttl_seconds  - entries older than this are refetched. Negative disables expiry.
    timeout_seconds    - per-request timeout. The default was 10s, which the 24-request
                         batch regularly exceeded, producing silently degraded reports.
    on_data            - called with the planning result as soon as it is known,
                         typically ~1-2s, without waiting for the map image. The
                         DP map export is roughly 95% of a cold lookup, so a caller
                         that reports from this callback shows an answer about ten
                         times sooner. metadata.documents_pending is True in that
                         snapshot; the returned dict is the final one. Exceptions
                         raised by the callback are caught and logged, never
                         allowed to abort the lookup.
    """
    try:
        safe_village = sanitize_query_value(village, _VILLAGE_RE, "village", 64)
        safe_cts = sanitize_query_value(cts_number, _CTS_RE, "cts_number", 32)
    except ValueError as exc:
        return {"error": str(exc), "suggestions": suggest_villages(village)}

    village_upper = safe_village.upper()
    cache_key = f"{village_upper}:{safe_cts}"

    os.makedirs(output_dir, exist_ok=True)

    if use_cache:
        t_cache = time.perf_counter()
        cached_res = read_cache_entry(output_dir, cache_key, cache_ttl_seconds)
        if cached_res is not None:
            # Report the real retrieval time. Previously this was hardcoded to 12.0,
            # and leaving it unset would replay the original cold run's duration -
            # either way the number would be fiction.
            cached_res["metadata"]["execution_time_ms"] = round((time.perf_counter() - t_cache) * 1000, 2)
            cached_res["metadata"]["cached_result"] = True
            age = cached_res["metadata"].get("cache_age_days", 0)
            print(
                f"[dp-lookup-pro] Serving cached report from "
                f"{cached_res['metadata'].get('cached_at')} ({age} days old). "
                f"Use --no-cache for a fresh check.",
                file=sys.stderr,
            )
            return cached_res

    t_start = time.perf_counter()
    register_path = os.path.join(output_dir, "dp-lookups.xlsx")

    # warnings = fetch failures. A run carrying any of these is never cached.
    # notes    = deterministic facts about the data itself (e.g. an area derived
    #            from geometry because MCGM has none on record). Informational
    #            only, and safe to cache.
    warnings: List[str] = []
    notes: List[str] = []

    limits = httpx.Limits(max_keepalive_connections=30, max_connections=50)
    async with httpx.AsyncClient(http2=True, timeout=timeout_seconds, limits=limits, follow_redirects=True) as client:
        query_params = {
            "where": f"VILLAGE='{village_upper}' AND CTS_CS_NO='{safe_cts}'",
            "outFields": "WARD,TYPE,VILLAGE,CTS_CS_NO,AREA_APP_SQ_MTRS,SHAPE.AREA",
            "returnGeometry": "true",
            "outSR": "102100",
            "f": "json",
        }
        try:
            resp = await client.get(f"{SERVER_URL}/13/query", params=query_params)
        except httpx.HTTPError as exc:
            # httpx timeout/connect errors often stringify to "", which tells the
            # user nothing. Always name the failure type.
            detail = str(exc).strip() or type(exc).__name__
            return {
                "error": (
                    f"Could not reach the MCGM map server ({detail}). "
                    "The server rate-limits sustained bursts - wait a few seconds and retry."
                )
            }

        data = usable_json(resp)
        if data is None:
            return {
                "error": (
                    "The MCGM map server rejected the parcel query "
                    f"(HTTP {getattr(resp, 'status_code', 'unknown')}). Try again shortly."
                )
            }

        if not data.get("features"):
            if village_upper in MCGM_VILLAGES:
                msg = (
                    f"Village '{village_upper}' is valid, but it has no plot "
                    f"numbered '{cts_number}'. Check the CTS number - suffixes and "
                    "slashes matter ('748A' and '748' are different plots, as are "
                    "'16/738' and '16-738')."
                )
            else:
                hints = suggest_villages(village_upper)
                msg = f"'{village}' is not a valid MCGM village name."
                if hints:
                    msg += " Did you mean: " + ", ".join(hints) + "?"
                msg += (" Village names are cadastral, not locality names - run "
                        "--list-villages to see all 128.")
            return {"error": msg, "suggestions": suggest_villages(village_upper)}

        feature = data["features"][0]
        attrs = feature["attributes"]
        rings = feature["geometry"]["rings"]

        area = resolve_plot_area(attrs)
        area_sqm, area_source = area["area_sqm"], area["area_source"]
        if area["note"]:
            notes.append(area["note"])
        attrs["AREA_APP_SQ_MTRS"] = area_sqm  # keep attrs consistent for renderers

        geom = prepare_geometry(rings)
        cx, cy = geom["cx"], geom["cy"]
        mcx, mcy = geom["mcx"], geom["mcy"]
        xs, ys, d = geom["xs"], geom["ys"], geom["d"]
        lat, lon = geom["lat"], geom["lon"]
        wgs_rings = geom["wgs_rings"]
        R = WEB_MERCATOR_R

        cts_clean = str(cts_number).replace('/', '-').replace('\\', '-')
        village_clean = village.lower().strip().replace('/', '-').replace('\\', '-').replace(' ', '_')
        query_dir = os.path.join(output_dir, bundle_folder_name(village, cts_number))
        os.makedirs(query_dir, exist_ok=True)

        # CRZ *zone polygons* — not boundary lines.
        # Previously this list held only line features (High Tide Line, Low Tide Line,
        # CRZ Lines & Boundaries, etc). A point-identify at a plot centroid can never
        # intersect a line, so the check could only ever return NO — a silent false
        # negative on every CRZ-affected plot.
        # Layer 14 carries the sub-tier in its `category` attribute (I / II / III / IV);
        # 1264 and 1548 corroborate. Layer 2238 ("Coastal Districts having CRZ") stays
        # excluded: it is district-wide and flags every plot in Mumbai as CRZ.
        crz_restriction_layer_ids = [14, 1264, 1548]

        ident_task = client.post(
            f"{SERVER_URL}/identify",
            data=identify_payload(cx, cy, planning_layers(crz_restriction_layer_ids),
                                  mcx, mcy, d, tolerance=30, return_geometry=False),
        )

        half = max(70, max(max(xs) - min(xs), max(ys) - min(ys)) * 0.9)
        W, H = 1000, 1000
        x0, x1 = mcx - half, mcx + half
        y0, y1 = mcy - half, mcy + half

        dp_snap_task = client.get(
            f"{SERVER_URL}/export",
            params=map_export_params(x0, y0, x1, y1, W, H),
        )

        road_sample_pts = road_probe_points(rings[0], mcx, mcy, xs, ys)
        road_tasks = [
            client.post(f"{SERVER_URL}/identify",
                        data=identify_payload(px, py, ROAD_LAYERS, mcx, mcy, d,
                                              tolerance=50, return_geometry=True))
            for px, py in road_sample_pts
        ]

        neighbor_sample_pts = neighbour_probe_points(mcx, mcy, d)
        neighbor_tasks = [
            client.post(f"{SERVER_URL}/identify",
                        data=identify_payload(px, py, NEIGHBOUR_LAYERS, mcx, mcy, d,
                                              tolerance=30, return_geometry=True))
            for px, py in neighbor_sample_pts
        ]

        zoom = 18
        xtile, ytile = latlon_to_tile(lat, lon, zoom)
        grid_dim = 3
        half_dim = grid_dim // 2
        sat_tile_tasks = []
        sat_coords = []
        for dy in range(-half_dim, half_dim + 1):
            for dx in range(-half_dim, half_dim + 1):
                tx, ty = xtile + dx, ytile + dy
                sat_tile_tasks.append(fetch_tile_cached(client, zoom, ty, tx))
                sat_coords.append((dx + half_dim, dy + half_dim))

        # Everything is dispatched at once, but the wait is split in two.
        #
        # Measured on MALABAR HILL 518: the /export map image alone took 12,334 ms
        # of a 13,008 ms lookup - 95% of the runtime - while every planning call
        # finished inside 500 ms. Awaiting the image work before reporting anything
        # made the whole answer as slow as its slowest picture.
        #
        # asyncio.gather() schedules immediately, so the slow half is genuinely
        # in flight while the fast half is being parsed.
        slow_group = asyncio.gather(
            dp_snap_task,
            asyncio.gather(*sat_tile_tasks, return_exceptions=True),
            return_exceptions=True,
        )

        fast_results = await asyncio.gather(
            ident_task,
            asyncio.gather(*road_tasks, return_exceptions=True),
            asyncio.gather(*neighbor_tasks, return_exceptions=True),
            return_exceptions=True
        )

        ident_resp = fast_results[0]
        road_resps = fast_results[1]
        neighbor_resps = fast_results[2]

        # Parse Roads
        road_map = {}
        road_geoms = []          # polylines in WGS84, for the DXF road layer
        road_failures = 0
        if isinstance(road_resps, Exception):
            road_failures = len(road_sample_pts)
        else:
            for r in road_resps:
                j = usable_json(r)
                if j is None:
                    road_failures += 1
                    continue
                for f in j.get('features', []):
                    r_attrs = f.get('attributes', {})
                    r_name = r_attrs.get('ROAD_NAME') or r_attrs.get('Roadname') or r_attrs.get('NAME') or r_attrs.get('Name') or r_attrs.get('TYPE_') or r_attrs.get('Type')
                    r_w = r_attrs.get('WIDTH_RL') or r_attrs.get('WIDTH') or r_attrs.get('Width')
                    if r_name or r_w:
                        road_map[f"{r_name}|{r_w}"] = {'name': r_name or 'Road', 'width': r_w or 'N/A'}
                for item in j.get('results', []):
                    r_attrs = item.get('attributes', {})
                    r_name = r_attrs.get('ROAD_NAME') or r_attrs.get('Roadname') or r_attrs.get('NAME') or r_attrs.get('Name') or r_attrs.get('TYPE_') or r_attrs.get('Type')
                    r_w = r_attrs.get('WIDTH_RL') or r_attrs.get('WIDTH') or r_attrs.get('Width')
                    if r_name or r_w:
                        road_map[f"{r_name}|{r_w}"] = {'name': r_name or 'Road', 'width': r_w or 'N/A'}
                    # Road layers are mixed geometry: 193/194 are polylines and
                    # return 'paths'; 44/45/2224 are road polygons and return 'rings'.
                    # Reading only 'paths' left C-ROAD-ALIGN empty on any plot
                    # whose frontage came from the polygon layers.
                    r_geom = item.get('geometry', {}) or {}
                    for part in (r_geom.get('paths') or []) + (r_geom.get('rings') or []):
                        if len(part) >= 2:
                            road_geoms.append([
                                [round((pt[0] / R) * 180, 7),
                                 round((math.atan(math.exp(pt[1] / R * math.pi)) * 2 - math.pi / 2) * 180 / math.pi, 7)]
                                for pt in part
                            ])

        if road_failures:
            warnings.append(
                f"{road_failures} of {len(road_sample_pts)} road probes failed; "
                "abutting road may be incomplete"
            )

        chosen_road = select_road(road_map)
        road_name, road_width = chosen_road["name"], chosen_road["width"]

        # Parse Neighbors
        neighbors_map = {}
        neighbor_failures = 0
        if isinstance(neighbor_resps, Exception):
            neighbor_failures = len(neighbor_sample_pts)
        else:
            for r in neighbor_resps:
                nj = usable_json(r)
                if nj is None:
                    neighbor_failures += 1
                    continue
                for item in nj.get('results', []):
                    n_attrs = item.get('attributes', {})
                    c_no = n_attrs.get('cts_cs_no') or n_attrs.get('CTS_CS_NO')
                    n_area = n_attrs.get('area_app_sq_mtrs') or n_attrs.get('AREA_APP_SQ_MTRS')
                    v_name = n_attrs.get('village') or n_attrs.get('VILLAGE')
                    n_geom = item.get('geometry', {}) or {}
                    # Identify returns Web Mercator (sr=102100). export_dxf expects
                    # WGS84 degrees like the main plot ring - without this the
                    # adjoining plots were drawn at ~8e11, far outside the sheet.
                    n_rings = [
                        [
                            [round((pt[0] / R) * 180, 7),
                             round((math.atan(math.exp(pt[1] / R * math.pi)) * 2 - math.pi / 2) * 180 / math.pi, 7)]
                            for pt in nr
                        ]
                        for nr in (n_geom.get('rings', []) or [])
                    ]
                    if c_no and str(c_no) != str(cts_number):
                        neighbors_map[c_no] = {
                            'cts_no': str(c_no),
                            'village': str(v_name),
                            'area_sqm': str(n_area) if n_area else 'N/A',
                            'rings': n_rings
                        }
        if neighbor_failures:
            warnings.append(
                f"{neighbor_failures} of {len(neighbor_sample_pts)} neighbour probes failed; "
                "adjoining plot list may be incomplete"
            )

        neighbors = list(neighbors_map.values())

        # Parse Identify.
        #
        # This call carries zone, reservation, designation, DP modification, metro
        # and CRZ - i.e. every planning remark in the report. If it FAILED we must
        # not continue: the defaults (zone 'Unknown', CRZ 'NO', no reservation) are
        # indistinguishable from a genuinely clear plot, and a full PDF docket built
        # on them looks authoritative. Fail loudly instead.
        ident_json = usable_json(ident_resp)
        if ident_json is None:
            return {
                "error": (
                    "Planning data could not be retrieved from the MCGM map server "
                    "(the identify request failed or timed out). No report was generated, "
                    "because a partial one would understate zoning, reservations and CRZ status. "
                    "Please retry."
                ),
                "partial_data": {
                    "village": village_upper,
                    "cts_no": str(cts_number),
                    "ward": str(attrs.get("WARD")),
                    "area_sqm": attrs.get("AREA_APP_SQ_MTRS"),
                },
            }

        results = ident_json.get("results", [])
        z_item = next((r for r in results if r.get("layerId") == 0), None)
        rv_item = next((r for r in results if r.get("layerId") == 46), None)
        des_item = next((r for r in results if r.get("layerId") == 47), None)
        mod_item = next((r for r in results if r.get("layerId") == 192), None)
        metro_item = next((r for r in results if r.get("layerId") == 1550), None)
        crz_item = next((r for r in results if r.get("layerId") in crz_restriction_layer_ids), None)

        zone = z_item["attributes"]["Zone_Code2"] if z_item else "Unknown"
        res_code = rv_item["attributes"]["NEW_RES_CODE_31"] if rv_item else "None"
        res_type = rv_item["attributes"]["NEW_RES_MAINTYPE_31"] if rv_item else "None"
        des_code = des_item["attributes"]["NEW_DES_CODE_31"] if des_item else "None"
        des_desc = des_item["attributes"]["DISCRIPTION"].strip() if des_item else "None"
        mod_label = mod_item["attributes"]["LABEL to Display"] if mod_item else "None"
        mod_approval = mod_item["attributes"]["APPROVAL_NO"] if mod_item else "None"
        mod_doc = mod_item["attributes"]["Approval Documents"] if mod_item else "None"
        
        metro_buffer_flag = "YES (Metro Buffer Zone)" if metro_item else "NO"

        crz_buffer_flag = crz_flag_from(crz_item)

        status = derive_status(mod_item, des_item, rv_item, des_desc, des_code,
                               res_code, res_type, mod_approval)
        status_badge, status_summary = status["badge"], status["summary"]

        ward_clean = str(attrs['WARD']).replace('/', '-').replace('\\', '-')

        # ------------------------------------------------------------------
        # FAST PATH COMPLETE. Everything below the on_data callback needs the
        # slow image fetches; everything above needed only planning data.
        # ------------------------------------------------------------------
        map_link = (
            f"https://mcgm.maps.arcgis.com/apps/webappviewer/index.html?"
            f"id=67118c3502fd492e94680d10e77ec112&marker={lon},{lat},4326,{cts_number}%20{village}"
            f"&center={lon},{lat}&level=19"
        )

        geojson_fname = f"plot_{ward_clean}_{cts_clean}_{village_clean}.geojson"
        geojson_path = os.path.join(query_dir, geojson_fname)
        dxf_fname = f"plot_{ward_clean}_{cts_clean}_{village_clean}.dxf"
        dxf_path = os.path.join(query_dir, dxf_fname)
        kml_fname = f"plot_{ward_clean}_{cts_clean}_{village_clean}.kml"
        kml_path = os.path.join(query_dir, kml_fname)
        dp_snapshot_fname = f"plot_{ward_clean}_{cts_clean}_{village_clean}_hd.png"
        dp_snapshot_path = os.path.join(query_dir, dp_snapshot_fname)
        sat_snapshot_fname = f"plot_{ward_clean}_{cts_clean}_{village_clean}_satellite.png"
        sat_snapshot_path = os.path.join(query_dir, sat_snapshot_fname)
        pdf_fname = f"dp_report_{ward_clean}_{cts_clean}_{village_clean}.pdf"
        pdf_path = os.path.join(query_dir, pdf_fname)

        export_props = {
            "village": village.upper(),
            "cts_no": str(cts_number),
            "ward": str(attrs['WARD']),
            "type": str(attrs['TYPE']),
            "area_sqm": attrs['AREA_APP_SQ_MTRS'],
            "zone": zone,
            "status_badge": status_badge,
            "reservation_code": res_code,
            "reservation_type": res_type,
            "designation_description": des_desc,
            "modification_approval": mod_approval,
            "crz_buffer_flag": crz_buffer_flag,
            "metro_buffer_flag": metro_buffer_flag,
            "abutting_road": road_name,
            "road_width": road_width,
            "area_source": area_source,
            "adjoining_cts_plots_count": len(neighbors),
            "map_link": map_link
        }

        # Vector exports need geometry and attributes only - no imagery - so they
        # are written now rather than behind the map fetch.
        export_geojson(wgs_rings, export_props, geojson_path)
        export_dxf(wgs_rings, export_props, dxf_path, neighbors=neighbors, roads=road_geoms[:6])
        export_kml(wgs_rings, export_props, kml_path)

        def _compose_result(exec_ms: float, pending: bool) -> Dict[str, Any]:
            return {
                "plot_identity": {
                    "village": village.upper(), "cts_no": str(cts_number),
                    "ward": str(attrs["WARD"]), "type": str(attrs["TYPE"]),
                    "area_sqm": area_sqm, "area_source": area_source,
                    "coordinates_wgs84": {"latitude": lat, "longitude": lon},
                },
                "planning_remarks": {
                    "status_badge": status_badge, "status_summary": status_summary, "zone": zone,
                    "reservation": {"code": res_code, "type": res_type},
                    "designation": {"code": des_code, "description": des_desc},
                    "dp_modification": {"approval_no": mod_approval, "details": mod_label,
                                        "document_link": mod_doc},
                },
                "regulatory_and_infrastructure": {
                    "crz_status": crz_buffer_flag, "metro_buffer": metro_buffer_flag,
                    "abutting_road": {"name": road_name, "width": road_width},
                },
                "spatial_cluster": {
                    "adjoining_plots_count": len(neighbors),
                    "adjoining_cts_plots": [{k: v for k, v in n.items() if k != 'rings'}
                                            for n in neighbors],
                },
                "export_files": {
                    "bundle_folder": query_dir, "pdf_report": pdf_path,
                    "hd_dp_map": dp_snapshot_path, "satellite_view": sat_snapshot_path,
                    "autocad_dxf": dxf_path, "autocad_geojson": geojson_path,
                    "google_earth_kml": kml_path, "master_excel_register": register_path,
                },
                "metadata": {
                    "source": "MCGM SDP 2014-34",
                    "lookup_datetime": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "execution_time_ms": exec_ms,
                    "cached_result": False,
                    "complete": not warnings,
                    "documents_pending": pending,
                    "warnings": list(warnings),
                    "notes": list(notes),
                    "interactive_web_map": map_link,
                },
            }

        if on_data is not None:
            fast_ms = round((time.perf_counter() - t_start) * 1000, 1)
            try:
                on_data(_compose_result(fast_ms, pending=True))
            except Exception as exc:
                print(f"[dp-lookup-pro] on_data callback raised: {exc}", file=sys.stderr)

        # ---- slow half: map image + satellite tiles, then PDF ----
        slow_results = await slow_group
        dp_snap_resp = slow_results[0]
        sat_tile_bytes = slow_results[1]

        # Render HD DP Map Overlay
        dp_map_ok = (
            not isinstance(dp_snap_resp, Exception)
            and getattr(dp_snap_resp, "status_code", None) == 200
            and dp_snap_resp.headers.get("content-type", "").startswith("image")
        )
        if not dp_map_ok:
            warnings.append("DP base map could not be fetched; the map image is a blank placeholder")

        final_dp_img = render_dp_map(
            dp_snap_resp.content if dp_map_ok else None,
            rings,
            bbox=(x0, y0, x1, y1),
            size=(W, H),
            labels={"cts": cts_number, "village": village.upper(), "ward": attrs["WARD"],
                    "zone": zone, "area": attrs["AREA_APP_SQ_MTRS"],
                    "status": status_badge[:30]},
        )

        # Stitch Satellite Canvas
        tiles = [] if isinstance(sat_tile_bytes, Exception) else list(sat_tile_bytes)
        missing_tiles = (len(sat_coords) if isinstance(sat_tile_bytes, Exception)
                         else sum(1 for b in tiles if isinstance(b, Exception) or not b))
        if missing_tiles:
            warnings.append(f"{missing_tiles} of {len(sat_coords)} satellite tiles failed to load")

        top_lat, left_lon = tile_to_latlon(xtile - half_dim, ytile - half_dim, zoom)
        bot_lat, right_lon = tile_to_latlon(xtile + half_dim + 1, ytile + half_dim + 1, zoom)

        final_sat = stitch_satellite(
            tiles or [None] * len(sat_coords),
            sat_coords,
            grid_dim,
            wgs_rings,
            bounds=(top_lat, left_lon, bot_lat, right_lon),
            labels={"cts": cts_number, "village": village.upper(), "ward": attrs["WARD"],
                    "lat": lat, "lon": lon, "neighbours": len(neighbors)},
        )

        def save_images():
            final_dp_img.convert("RGB").save(dp_snapshot_path, "PNG", compress_level=1)
            final_sat.convert("RGB").save(sat_snapshot_path, "PNG", compress_level=1)

        await asyncio.to_thread(save_images)

        qr = qrcode.QRCode(box_size=4, border=1)
        qr.add_data(map_link)
        qr.make(fit=True)
        qr_img = qr.make_image(fill_color="black", back_color="white")
        qr_bytes = io.BytesIO()
        qr_img.save(qr_bytes, format="PNG")
        qr_bytes.seek(0)
        
        await asyncio.to_thread(
            build_pdf_doc,
            pdf_path, status_badge, status_summary, village, attrs, cts_number, zone, des_desc, des_code, mod_approval, mod_label, crz_buffer_flag, metro_buffer_flag, road_name, road_width, dp_snapshot_path, qr_bytes, map_link, sat_snapshot_path, neighbors,
            ("Approved cadastral area (MCGM record)"
             if area_source.startswith("approved")
             else "DERIVED from plot boundary - no approved area on MCGM record")
        )

        now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        headers = [
            "lookup_datetime", "source", "ward", "village", "cts_no", "type",
            "area_sqm", "zone", "status_badge", "reservation_code", "reservation_type",
            "designation_code", "designation_desc", "modification_approval",
            "modification_label", "modification_doc", "crz_buffer", "metro_buffer",
            "abutting_road", "road_width", "adjoining_plots_count", "hd_snapshot_file",
            "satellite_snapshot_file", "pdf_report_file", "dxf_file", "geojson_file", "kml_file", "map_link"
        ]
        row = [
            now_str, "MCGM SDP 2014-34", attrs["WARD"], village, cts_number, attrs["TYPE"],
            attrs["AREA_APP_SQ_MTRS"], zone, status_badge, res_code, res_type,
            des_code, des_desc, mod_approval, mod_label, mod_doc, crz_buffer_flag, metro_buffer_flag,
            road_name, road_width, len(neighbors), dp_snapshot_fname, sat_snapshot_fname, pdf_fname,
            dxf_fname, geojson_fname, kml_fname, map_link
        ]

        wb = load_workbook(register_path) if os.path.exists(register_path) else Workbook()
        ws = wb.active
        if ws.max_row == 1 and ws["A1"].value is None:
            ws.append(headers)
        ws.append(row)
        wb.save(register_path)

        t_end = time.perf_counter()
        exec_ms = round((t_end - t_start) * 1000, 1)

        result_dict = _compose_result(exec_ms, pending=False)

        if warnings:
            print(f"[dp-lookup-pro] WARNING: result incomplete, not cached - {'; '.join(warnings)}",
                  file=sys.stderr)
        else:
            write_cache_entry(output_dir, cache_key, result_dict)

        return result_dict


# The 128 valid MCGM revenue village names, read from layer 13 on 2026-07-28.
# These are cadastral survey names and are effectively static - layer 13 has not
# been edited since 2019-01-23. Kept local so name help costs no network call.
# Refresh with:
#   GET {SERVER_URL}/13/query?where=1%3D1&outFields=VILLAGE
#       &returnDistinctValues=true&returnGeometry=false&f=json
MCGM_VILLAGES = (
    "AAKSE", "AAREY", "AKURLI", "AMBIVALI", "ANDHERI", "ANIK", "ASALPE", "BANDIVALI",
    "BANDRA-A", "BANDRA-B", "BANDRA-C", "BANDRA-D", "BANDRA-E", "BANDRA-EAST",
    "BANDRA-F", "BANDRA-G", "BANDRA-H", "BANDRA-I", "BAPNALA", "BHANDUP-E", "BHANDUP-W",
    "BHULESHWAR", "BORIVALI", "BORLA", "BRAMHANWADA", "BYCULLA", "CHAKALA", "CHANDIVALI",
    "CHARKOP", "CHEMBUR", "CHINCHAVALI", "COLABA", "DADAR-NAIGAON", "DAHISAR", "DARAVALI",
    "DEONAR", "DHARAVI", "DINDOSHI", "EKSAR", "ERANGAL", "FORT", "GHATKOPAR",
    "GHATKOPAR KIROL", "GIRGAUM", "GORAI", "GOREGAON", "GUNDAVALI", "GUNDHGAON",
    "HARIYALI-E", "HARIYALI-W", "ISMALIA", "JUHU", "KANDIVALI", "KANHERI", "KANJUR-E",
    "KANJUR-W", "KIROL", "KLERABAD", "KOLEKALYAN", "KOLEKALYAN UNIVERSITY", "KONDIVATE",
    "KOPRI", "KURAR", "KURLA - 1", "KURLA - 2", "KURLA - 3", "KURLA - 4", "LOWER PAREL",
    "MADH", "MAGATHANE", "MAHIM", "MAHUL", "MAJAS", "MALABAR HILL", "MALAD", "MALAD-E",
    "MALAD-NORTH", "MALAD-SOUTH", "MALVANI", "MANDALE", "MANDPESHWAR-M", "MANDPESHWAR-N",
    "MANDPESHWAR-S", "MANDVI", "MANKHURD", "MANORI", "MARAVALI", "MAROL", "MAROL MAROSHI",
    "MARVE", "MATUNGA", "MAZAGAON", "MOGRA", "MOHILI", "MULGAON", "MULUND-E", "MULUND-W",
    "NAHUR", "OSHIWARA", "PAHADI EKSAR", "PAHADI GOREGAON-E", "PAHADI GOREGAON-W",
    "PAREL-SEWERI", "PARIGHIKARI", "PASPOLI", "POISAR", "POWAI", "PRAJAPUR",
    "PRINCESS DOCK", "SAAI", "SAHAR", "SAKI", "SALT PAN", "SHIMPAWALI", "SION", "TARDEO",
    "TIRANDAZ", "TULSI", "TUNGWE", "TURBHE", "VALNAI", "VERSOVA", "VIKHROLI", "VILE PARLE",
    "VYARAVLI", "WADHAVALI", "WADHWAN", "WORLI",
)


def suggest_villages(name: Any, limit: int = 8) -> List[str]:
    """
    Nearest valid village names for a mistyped one.

    The commonest failure by far is a modern locality name that is not a
    cadastral village: BANDRA is not valid (it is BANDRA-A..BANDRA-I plus
    BANDRA-EAST), KURLA is not valid (KURLA - 1..4), BHANDUP is not valid
    (BHANDUP-E/W). Prefix matches are checked before fuzzy ones so those
    families surface first.
    """
    import difflib

    query = str(name or "").strip().upper()
    if not query:
        return []
    hits = [v for v in MCGM_VILLAGES if v == query]
    hits += [v for v in MCGM_VILLAGES if v.startswith(query) and v not in hits]
    hits += [v for v in MCGM_VILLAGES if query in v and v not in hits]
    for v in difflib.get_close_matches(query, MCGM_VILLAGES, n=limit, cutoff=0.6):
        if v not in hits:
            hits.append(v)
    return hits[:limit]


def format_result_human(result: Dict[str, Any]) -> str:
    """Readable summary of a lookup. Raw JSON stays available behind --json."""
    if "error" in result:
        lines = ["", "  Could not complete the lookup:", f"  {result['error']}", ""]
        return "\n".join(lines)

    ident = result["plot_identity"]
    plan = result["planning_remarks"]
    reg = result["regulatory_and_infrastructure"]
    cluster = result["spatial_cluster"]
    files = result["export_files"]
    meta = result["metadata"]

    area = ident.get("area_sqm")
    area_txt = f"{area:,.2f} m\u00b2" if isinstance(area, (int, float)) else "not on record"
    if "derived" in str(ident.get("area_source", "")):
        area_txt += "  (derived from boundary, not MCGM-approved)"

    rows = [
        ("Plot area", area_txt),
        ("Zone", plan.get("zone")),
        ("Reservation", plan["reservation"].get("type") if plan["reservation"].get("code") != "None" else "None"),
        ("Designation", plan["designation"].get("description")),
        ("DP modification", plan["dp_modification"].get("approval_no")),
        ("CRZ", reg.get("crz_status")),
        ("Metro buffer", reg.get("metro_buffer")),
        ("Abutting road", f"{reg['abutting_road'].get('name')} ({reg['abutting_road'].get('width')})"),
        ("Adjoining plots", str(cluster.get("adjoining_plots_count"))),
    ]

    out = ["", f"  {plan.get('status_badge')}",
           f"  {ident.get('village')}  ·  CTS {ident.get('cts_no')}  ·  Ward {ident.get('ward')}", ""]
    for label, value in rows:
        out.append(f"    {label:<17} {value}")

    neighbours = [n.get("cts_no") for n in cluster.get("adjoining_cts_plots", [])][:6]
    if neighbours:
        out.append(f"    {'Neighbours':<17} CTS " + ", ".join(str(n) for n in neighbours))

    out += ["", f"    Files            {files.get('bundle_folder')}  (6 files)",
            f"    Report           {files.get('pdf_report')}",
            f"    Register         {files.get('master_excel_register')}"]

    if meta.get("cached_result"):
        out.append(f"\n    Cached result from {meta.get('cached_at')} "
                   f"({meta.get('cache_age_days')} days old). Use --no-cache to refetch.")
    else:
        out.append(f"\n    Fetched in {meta.get('execution_time_ms', 0) / 1000:.1f}s")

    for note in meta.get("notes") or []:
        out.append(f"    Note: {note}")
    if not meta.get("complete", True):
        out.append("\n    INCOMPLETE - this report is missing data:")
        for w in meta.get("warnings") or []:
            out.append(f"      - {w}")
        out.append("    Re-run before relying on it.")

    out.append("")
    return "\n".join(out)


USAGE = """Usage: dp-lookup-pro <VILLAGE_NAME> <CTS_NUMBER> [OUTPUT_DIR] [options]

Options:
  --json            print the full JSON response instead of a summary
  --no-cache        force a fresh lookup, ignoring any cached report
  --list-villages   print all 128 valid village names and exit

Examples:
  dp-lookup-pro WORLI 947
  dp-lookup-pro "MALABAR HILL" "16/738"
  dp-lookup-pro BANDRA-A 409 ./client-reports --no-cache

Village must be one of the 128 exact MCGM revenue village names
(e.g. BANDRA-A, not BANDRA). See START-HERE.md for the full list."""


def main(argv: Optional[List[str]] = None) -> int:
    """Console entry point. Returns a process exit code."""
    args = list(sys.argv[1:] if argv is None else argv)

    if "--list-villages" in args or "--villages" in args:
        print(f"\n{len(MCGM_VILLAGES)} valid MCGM village names:\n")
        for i in range(0, len(MCGM_VILLAGES), 3):
            print("  " + "".join(f"{v:<26}" for v in MCGM_VILLAGES[i:i + 3]).rstrip())
        print("\nNames are exact. BANDRA, KURLA and BHANDUP alone are NOT valid -")
        print("use BANDRA-A..BANDRA-I, KURLA - 1..4, BHANDUP-E/W.\n")
        return 0

    as_json = False
    for flag in ("--json", "--raw"):
        if flag in args:
            args.remove(flag)
            as_json = True

    use_cache = True
    for flag in ("--no-cache", "--fresh"):
        if flag in args:
            args.remove(flag)
            use_cache = False

    # Tolerate the conversational form: "WORLI CTS 947"
    if len(args) >= 3 and args[1].upper() in ("CTS", "CS", "PLOT", "NO", "NO."):
        args = [args[0], args[2], *args[3:]]

    if len(args) < 2:
        print(USAGE)
        return 1

    village, cts_number = args[0], args[1]
    output_dir = args[2] if len(args) > 2 else "./output"

    if not as_json:
        print(f"  Looking up {village} CTS {cts_number}"
              f"{' (fresh)' if not use_cache else ''}...")
    # Print the planning answer the moment it is known, then let the map image,
    # satellite tiles and PDF finish. The user sees the result in ~1-2s instead
    # of waiting on an export call that is ~95% of the runtime.
    shown = {"done": False}

    def _show_early(snapshot):
        if as_json or shown["done"]:
            return
        shown["done"] = True
        print(format_result_human(snapshot))
        print("  Building PDF, maps and CAD files...", flush=True)

    result = asyncio.run(
        lookup_plot_pro(
            village=village,
            cts_number=cts_number,
            output_dir=output_dir,
            use_cache=use_cache,
            on_data=None if as_json else _show_early,
        )
    )

    if as_json:
        print(json.dumps(result, indent=2, ensure_ascii=False))
    elif shown["done"] and "error" not in result:
        meta = result.get("metadata", {})
        print(f"  Done in {meta.get('execution_time_ms', 0) / 1000:.1f}s - all 6 files written.")
        for w in meta.get("warnings") or []:
            print(f"  Warning: {w}")
        print()
    else:
        print(format_result_human(result))
    return 1 if "error" in result else 0


if __name__ == "__main__":
    sys.exit(main())
